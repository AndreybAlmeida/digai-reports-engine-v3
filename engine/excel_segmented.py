"""
DigAI Reports Engine — Excel Segmentado

Gera Excel com múltiplos segmentos:
  Aba 1: Sumário Geral (tabela comparativa com ranking)
  Aba 2..N: Cada segmento (layout idêntico ao relatório padrão)
  Aba Final: Base de Dados completa

Todas as fórmulas das abas de segmento usam COUNTIFS com critério
adicional do segmento, apontando para a aba Base de Dados.
"""

from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .excel_gen import (
    _fill, _font, _align, _border_thin,
    _build_base,
    C_NAVY, C_BLUE, C_ACCENT, C_TEAL, C_WHITE,
    C_LIGHT_BG, C_HEADER_BG, C_BORDER, C_YELLOW, FONT_DEFAULT,
)
from .dimensions import build_summary_table


def _slug_sheet(name: str) -> str:
    """Converte nome em título de aba Excel (máx 31 chars, sem chars inválidos)."""
    import re
    s = re.sub(r"[/\\*\[\]:?]", "", str(name))
    return s[:31]


def _build_premissas_openpyxl(wb: Workbook, params: dict) -> None:
    """Aba Premissas — openpyxl (usada no relatório segmentado)."""
    ws = wb.create_sheet("Premissas")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "Premissas e Parâmetros"
    t.font = _font(bold=True, size=14, color=C_WHITE)
    t.fill = _fill(C_NAVY)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 32

    for j, h in enumerate(["Parâmetro", "Valor", "Descrição"], start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = _font(bold=True, size=10, color=C_WHITE)
        c.fill = _fill(C_ACCENT)
        c.alignment = _align("center")
        c.border = _border_thin()

    params_table = [
        ("Cliente",               params.get("cliente_nome", "Cliente"),        "@",            "Nome do cliente para o relatório"),
        ("Período",               params.get("periodo", ""),                    "@",            "Período de referência do relatório"),
        ("Mensalidade DigAI",     params.get("mensalidade_digai", 7600.0),      "R$ #,##0.00",  "Valor mensal pago ao DigAI (R$)"),
        ("Salário TA CLT",        params.get("salario_ta_clt", 4750.0),         "R$ #,##0.00",  "Salário bruto do Talent Acquisition CLT (R$)"),
        ("Horas por Mês (TA)",    params.get("horas_mes", 176),                 "#,##0",        "Horas mensais trabalhadas pelo TA (22 dias x 8h)"),
        ("Produtividade (%)",     params.get("produtividade_pct", 0.60),        "0%",           "% do tempo do TA dedicada a entrevistas"),
        ("Duração da EI (min)",   params.get("tempo_entrevista_min", 30),       "#,##0",        "Duração média de uma entrevista presencial (min)"),
        ("Capacidade Max. TA/mês",params.get("max_entrevistas_ta", 127),        "#,##0",        "Max. entrevistas que o TA consegue fazer por mês"),
    ]

    for i, (label, value, fmt, desc) in enumerate(params_table):
        r = 4 + i
        ws.row_dimensions[r].height = 22
        alt = i % 2 == 0
        bg_lbl = C_LIGHT_BG if alt else "FFFFFF"

        lc = ws.cell(row=r, column=1, value=label)
        lc.font = _font(size=10)
        lc.fill = _fill(bg_lbl)
        lc.alignment = _align("left")
        lc.border = _border_thin()

        # Valor editável → fundo amarelo
        vc = ws.cell(row=r, column=2, value=value)
        vc.font = _font(bold=True, size=10, color="#000000")
        vc.fill = _fill(C_YELLOW)
        vc.alignment = _align("right")
        vc.border = _border_thin()
        if fmt and fmt != "@":
            vc.number_format = fmt

        dc = ws.cell(row=r, column=3, value=desc)
        dc.font = _font(size=9, italic=True, color="64748B")
        dc.fill = _fill(C_LIGHT_BG if alt else "FFFFFF")
        dc.alignment = _align("left")
        dc.border = _border_thin()

    note_r = 4 + len(params_table) + 1
    ws.merge_cells(f"A{note_r}:C{note_r}")
    nc = ws[f"A{note_r}"]
    nc.value = "Edite os valores na coluna B. Todas as outras abas atualizam automaticamente."
    nc.font = _font(size=9, italic=True, color="64748B")
    nc.alignment = _align("left")


def _build_sumario(wb: Workbook, summary_rows: list, results: dict):
    """
    Aba Sumário Geral — tabela comparativa por segmento (REGRA 4).

    Colunas: Dimensão | Total Cand. DigAI | Adesão % | Assertividade % |
             Contratações DigAI | SLA Médio | Volume DigAI % | Perf. vs Média
    """
    ws = wb.create_sheet("Sumário Geral")
    ws.sheet_view.zoomScale = 100
    ws.sheet_view.showGridLines = False

    # Larguras
    ws.column_dimensions["A"].width = 28
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 17

    # Título
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "Sumário Geral — Comparativo por Dimensão"
    t.font = _font(bold=True, size=13, color=C_WHITE)
    t.fill = _fill(C_NAVY)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    headers = [
        "Dimensão", "Total Cand. DigAI", "Adesão %", "Assertividade %",
        "Contratações DigAI", "SLA Médio", "Volume DigAI %", "Perf. vs Média",
    ]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = _font(bold=True, size=10, color=C_WHITE)
        c.fill = _fill(C_BLUE)
        c.alignment = _align("center")
        c.border = _border_thin()

    # Identifica melhor e pior segmento por assertividade
    valid_assert = [(i, r["assertividade"]) for i, r in enumerate(summary_rows)
                    if r.get("assertividade") is not None]
    best_idx = max(valid_assert, key=lambda x: x[1])[0] if valid_assert else -1
    worst_idx = min(valid_assert, key=lambda x: x[1])[0] if valid_assert else -1

    # Totais para linha de rodapé
    tot_com  = sum(r["total_com"] for r in summary_rows)
    tot_cont = sum(r["contratados"] for r in summary_rows)
    avg_ades = (sum(r["adesao"] or 0 for r in summary_rows) / len(summary_rows)
                if summary_rows else 0)
    avg_asrt = (sum(r["assertividade"] or 0 for r in summary_rows) / len(summary_rows)
                if summary_rows else 0)
    sla_vals = [r["sla_media"] for r in summary_rows if r.get("sla_media") is not None]
    avg_sla  = sum(sla_vals) / len(sla_vals) if sla_vals else None
    avg_vol  = (sum(r.get("volume_pct", 0) for r in summary_rows) / len(summary_rows)
                if summary_rows else 0)

    col_fmts = [None, "#,##0", "0.0%", "0.0%", "#,##0", "#,##0.0", "0.0%", "+0.0%;-0.0%;0.0%"]

    for i, row in enumerate(summary_rows):
        r = 3 + i
        alt = i % 2 == 1
        if i == best_idx:
            bg = "#DCFCE7"   # verde claro — melhor segmento
        elif i == worst_idx and best_idx != worst_idx:
            bg = "#FEE2E2"   # vermelho claro — pior segmento
        else:
            bg = C_LIGHT_BG if alt else "FFFFFF"

        perf = row.get("perf_vs_media")
        values = [
            row["segmento"],
            row["total_com"],
            (row["adesao"] or 0) / 100 if isinstance(row["adesao"], (int, float)) else row["adesao"],
            (row["assertividade"] or 0) / 100 if isinstance(row["assertividade"], (int, float)) else row["assertividade"],
            row["contratados"],
            row["sla_media"],
            row.get("volume_pct", 0),
            (perf / 100) if isinstance(perf, (int, float)) else perf,
        ]

        for j, (val, fmt) in enumerate(zip(values, col_fmts), start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.font = _font(size=10, bold=(i == best_idx))
            c.fill = _fill(bg)
            c.alignment = _align("right" if j > 1 else "left")
            c.border = _border_thin()
            if fmt and val is not None:
                c.number_format = fmt

    # Linha de totais/médias em navy
    tot_row = 3 + len(summary_rows)
    ws.row_dimensions[tot_row].height = 20
    tot_values = [
        "TOTAL / MÉDIA",
        tot_com,
        avg_ades / 100 if avg_ades else 0,
        avg_asrt / 100 if avg_asrt else 0,
        tot_cont,
        avg_sla,
        avg_vol,
        0,
    ]
    for j, (val, fmt) in enumerate(zip(tot_values, col_fmts), start=1):
        c = ws.cell(row=tot_row, column=j, value=val)
        c.font = _font(bold=True, size=10, color=C_WHITE)
        c.fill = _fill(C_NAVY)
        c.alignment = _align("right" if j > 1 else "left")
        c.border = _border_thin()
        if fmt and val is not None:
            c.number_format = fmt

    # Legenda
    note_row = tot_row + 2
    ws.merge_cells(f"A{note_row}:H{note_row}")
    nc = ws[f"A{note_row}"]
    nc.value = (
        f"Verde = melhor segmento  |  Vermelho = pior segmento  |  "
        f"{len(summary_rows)} segmentos  |  Ordenado por Assertividade % desc."
    )
    nc.font = _font(size=9, italic=True, color="64748B")
    nc.alignment = _align("left")


def _build_segment_sheet(wb: Workbook, seg_name: str, relatorio: dict,
                         dim_col_excel: str, dim_col_label: str):
    """
    Cria aba para um segmento específico.
    Fórmulas usam COUNTIFS com critério adicional: BaseDados[coluna]=seg_name
    """
    ws = wb.create_sheet(_slug_sheet(seg_name))
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18

    # Critério de filtro a ser adicionado em todos os COUNTIFS
    seg_filter = f',BaseDados[{dim_col_label}],"{seg_name}"'

    kpis = relatorio["kpis"]
    roi  = relatorio["roi"]
    ins  = relatorio["insights"]
    com  = kpis.get("Com DigAI", {})
    sem  = kpis.get("Sem DigAI", {})

    # Cabeçalho do segmento
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = f"Segmento: {seg_name}  ·  {relatorio['meta']['cliente']}  ·  {relatorio['meta']['periodo']}"
    t.font = _font(bold=True, size=12, color=C_WHITE)
    t.fill = _fill(C_NAVY)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value = f"Dimensão: {relatorio['meta'].get('dimensao','')}  ·  Filtro: {dim_col_label} = {seg_name}"
    sub.font = _font(size=9, italic=True, color="64748B")
    sub.alignment = _align("right")

    # KPIs com fórmulas filtradas por segmento
    section_row = 4
    ws.merge_cells(f"A{section_row}:D{section_row}")
    sc = ws[f"A{section_row}"]
    sc.value = "KPIs do Segmento"
    sc.font = _font(bold=True, size=10, color=C_ACCENT)
    sc.fill = _fill(C_HEADER_BG)
    sc.alignment = _align("left")

    headers = ["Indicador", "Com DigAI", "Sem DigAI", "Δ"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=section_row + 1, column=j, value=h)
        c.font = _font(bold=True, size=9, color=C_WHITE)
        c.fill = _fill(C_BLUE)
        c.alignment = _align("center")
        c.border = _border_thin()

    kpi_rows = [
        ("Total candidatos",
         f'=COUNTIFS(BaseDados[Processo Seletivo],"Com DigAI"{seg_filter})',
         f'=COUNTIFS(BaseDados[Processo Seletivo],"Sem DigAI"{seg_filter})',
         "#,##0"),
        ("Contratações",
         f'=COUNTIFS(BaseDados[Processo Seletivo],"Com DigAI",BaseDados[Status],"Contratado"{seg_filter})',
         f'=COUNTIFS(BaseDados[Processo Seletivo],"Sem DigAI",BaseDados[Status],"Contratado"{seg_filter})',
         "#,##0"),
        ("Adesão",
         f'=IFERROR(1-COUNTIFS(BaseDados[Processo Seletivo],"Com DigAI",BaseDados[Status],"Desistiu",BaseDados[Data Entrevista DigAI],"<>"{seg_filter})/COUNTIFS(BaseDados[Processo Seletivo],"Com DigAI",BaseDados[Data Entrevista DigAI],"<>"{seg_filter}),0)',
         '"N/A"',
         "0.0%"),
        ("SLA médio (dias)",
         f'=IFERROR(AVERAGEIFS(BaseDados[Data Final]-BaseDados[Data Cadastro],BaseDados[Processo Seletivo],"Com DigAI",BaseDados[Status],"Contratado"{seg_filter}),0)',
         f'=IFERROR(AVERAGEIFS(BaseDados[Data Final]-BaseDados[Data Cadastro],BaseDados[Processo Seletivo],"Sem DigAI",BaseDados[Status],"Contratado"{seg_filter}),0)',
         "#,##0.0"),
    ]

    for i, (label, f_com, f_sem, fmt) in enumerate(kpi_rows):
        r = section_row + 2 + i
        alt = i % 2 == 1
        bg = C_LIGHT_BG if alt else "FFFFFF"

        for col, val in [(1, label), (2, f_com), (3, f_sem)]:
            c = ws.cell(row=r, column=col, value=val)
            c.font = _font(size=10, bold=(col == 1))
            c.fill = _fill(bg)
            c.alignment = _align("left" if col == 1 else "right")
            c.number_format = fmt if col > 1 else "@"
            c.border = _border_thin()

        # Delta
        dc = ws.cell(row=r, column=4, value=f"=B{r}-C{r}" if "N/A" not in f_sem else "")
        dc.font = _font(size=10)
        dc.fill = _fill(bg)
        dc.alignment = _align("right")
        dc.number_format = "+#,##0.0;-#,##0.0;0" if "%" not in fmt else "+0.0%;-0.0%;0.0%"
        dc.border = _border_thin()

    # Veredicto do segmento
    verd_r = section_row + 2 + len(kpi_rows) + 1
    ws.merge_cells(f"A{verd_r}:D{verd_r}")
    vc = ws[f"A{verd_r}"]
    vc.value = f"Veredicto: {ins['veredicto']}"
    vc.font = _font(bold=True, size=11, color=C_WHITE)
    vc.fill = _fill("166534" if "BEM" in ins["veredicto"] else
                    "92400E" if "ATENÇÃO" in ins["veredicto"] else "991B1B")
    vc.alignment = _align("center")
    ws.row_dimensions[verd_r].height = 24

    # Funil
    funil_start = verd_r + 2
    ws.merge_cells(f"A{funil_start}:D{funil_start}")
    fsc = ws[f"A{funil_start}"]
    fsc.value = "Funil de Conversão"
    fsc.font = _font(bold=True, size=10, color=C_ACCENT)
    fsc.fill = _fill(C_HEADER_BG)
    fsc.alignment = _align("left")

    funil = relatorio.get("funil_din") or relatorio.get("funil", [])
    for i, row_d in enumerate(funil):
        r = funil_start + 1 + i
        alt = i % 2 == 1
        bg = C_LIGHT_BG if alt else "FFFFFF"
        for col, val, fmt in [
            (1, row_d["etapa"], None),
            (2, row_d["com_digai"], "#,##0"),
            (3, row_d["sem_digai"], "#,##0"),
            (4, row_d["pct_com"] / 100 if isinstance(row_d["pct_com"], (int, float)) else 0, "0.0%"),
        ]:
            c = ws.cell(row=r, column=col, value=val)
            c.font = _font(size=9)
            c.fill = _fill(bg)
            c.alignment = _align("left" if col == 1 else "right")
            c.border = _border_thin()
            if fmt:
                c.number_format = fmt


def gerar_excel_segmentado(
    results: dict,       # {seg_name: relatorio_dict}
    config: dict,
    params: dict,
    output_path: str,
) -> str:
    """
    Gera Excel com abas por segmento + aba Sumário + Base de Dados.
    """
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # ── Aba Base de Dados (dados de todos os segmentos juntos) ─────────────────
    # Reúne todos os _df
    import pandas as pd
    dfs = [rel["_df"] for rel in results.values() if rel.get("_df") is not None]
    if dfs:
        df_all = pd.concat(dfs, ignore_index=True)
        df_all.attrs.update(next(iter(results.values()))["_df"].attrs)
        print("   📋 Construindo Base de Dados...")
        col_map = _build_base(wb, df_all)
    else:
        col_map = {}

    # ── Aba Sumário Geral ──────────────────────────────────────────────────────
    summary_rows = build_summary_table(results)
    print("   📊 Construindo Sumário Geral...")
    _build_sumario(wb, summary_rows, results)

    # ── Abas por segmento ──────────────────────────────────────────────────────
    dim_col_label = config.get("dimension", "Segmento").capitalize()

    for seg_name, relatorio in results.items():
        print(f"   📁 Aba: {seg_name[:20]}")
        _build_segment_sheet(
            wb=wb,
            seg_name=seg_name,
            relatorio=relatorio,
            dim_col_excel=config.get("dim_col", ""),
            dim_col_label=dim_col_label,
        )

    # ── Aba Premissas ──────────────────────────────────────────────────────────
    print("   ⚙️  Premissas...")
    _build_premissas_openpyxl(wb, params)

    # Reordena: Sumário → Segmentos → Premissas → Base
    desired_start = ["Sumário Geral"]
    desired_end   = ["Premissas", "Base de Dados"]
    sheets = {ws.title: i for i, ws in enumerate(wb.worksheets)}
    for i, title in enumerate(desired_start):
        if title in sheets:
            wb.move_sheet(title, offset=-sheets[title])

    wb.save(output_path)
    print(f"   ✅ Excel segmentado: {output_path}")
    return output_path
