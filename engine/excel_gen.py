"""
DigAI Reports Engine — Excel Generator (LAYOUT_SPEC.md v1.0)

Abas geradas (ordem obrigatória per spec §4):
  1. Indicadores DigAI   — KPIs, Saving, ROI + definições
  2. Calculadora ROI     — parâmetros editáveis (amarelo) + resultados calculados
  3. Score & Qualidade   — apenas se score_ia disponível
  4. Evolução Mensal     — contratações mês a mês
  5. Por Departamento    — apenas se coluna de área/departamento disponível
  6. Base de Dados       — raw auditável (xlsxwriter constant_memory, row-by-row)

Todas as células de entrada têm fundo #FFF2CC (AMARELO) e borda #FFD966.
Todas as células calculadas "Com DigAI" têm fundo #BDD7EE (AZUL_CLARO).
show_gridlines = False em todas as abas.
"""

import gc
import re
import math
from pathlib import Path
from datetime import datetime, date

import pandas as pd
import numpy as np
import xlsxwriter


# ─── Paleta canonical (LAYOUT_SPEC.md §1) ────────────────────────────────────
C_AZUL_ESCURO   = "#1F3864"   # Headers principais, títulos, sidebar
C_AZUL_MEDIO    = "#2E75B6"   # Headers secundários, sub-títulos de bloco
C_AZUL_CLARO    = "#BDD7EE"   # Células calculadas "Com DigAI"
C_VERDE_ESCURO  = "#375623"   # Texto resultado positivo
C_VERDE_CLARO   = "#E2EFDA"   # Fundo resultado positivo
C_AMARELO       = "#FFF2CC"   # Células editáveis (input), referência Sem DigAI
C_AMARELO_BORDA = "#FFD966"   # Borda células editáveis
C_CINZA_CLARO   = "#F2F2F2"   # Linhas alternadas (zebra)
C_BRANCO        = "#FFFFFF"   # Fundo padrão
C_VERMELHO      = "#C00000"   # Alertas, reprovados
C_VERMELHO_CLARO= "#FFE7E7"   # Fundo alertas/reprovados
C_LARANJA       = "#C55A11"   # Desistentes, avisos
C_LARANJA_CLARO = "#FCE4D6"   # Fundo desistentes/avisos
C_CINZA_NA      = "#D9D9D9"   # Células KPI incalculável (N/A)
C_BORDER_STD    = "#9DC3E6"   # Borda padrão thin
C_BORDER_THICK  = "#1F3864"   # Borda medium (títulos principais)

# ─── Aliases de compatibilidade (excel_segmented.py e código legado) ─────────
C_NAVY      = C_AZUL_ESCURO
C_BLUE      = C_AZUL_MEDIO
C_ACCENT    = C_AZUL_MEDIO
C_TEAL      = C_AZUL_CLARO
C_WHITE     = C_BRANCO
C_LIGHT     = C_CINZA_CLARO
C_HEADER    = C_CINZA_CLARO
C_BORDER    = C_BORDER_STD
C_LIGHT_BG  = C_CINZA_CLARO
C_HEADER_BG = C_CINZA_CLARO
C_YELLOW    = C_AMARELO
C_POSITIVE  = C_VERDE_ESCURO
C_POS_BG    = C_VERDE_CLARO
C_NEGATIVE  = C_VERMELHO
C_NEG_BG    = C_VERMELHO_CLARO
C_WARNING   = C_LARANJA
C_WARN_BG   = C_LARANJA_CLARO

FONT_DEFAULT = "Calibri"
FONT_TITLE   = "Calibri"

# ─── Nomes de abas (constantes para cross-sheet refs) ─────────────────────────
_INDICADORES_SHEET  = "Indicadores DigAI"
_ROI_SHEET          = "Calculadora ROI"
_POR_TIME_SHEET     = "Por Time"
_PIVOT_SHEET        = "Pivot Calculos"
_BASE_SHEET         = "Base de Dados"
# Mantidos para compatibilidade mas nao adicionados ao workbook por padrao
_SCORE_SHEET        = "Score & Qualidade"
_MENSAL_SHEET       = "Evolucao Mensal"
_DEPTO_SHEET        = "Por Departamento"

# Referências de células do _ROI_SHEET (1-indexed, col C = idx 2)
# Linha: o layout de _build_calculadora_roi define as linhas abaixo
_ROI = {
    "salario_clt":    f"'{_ROI_SHEET}'!C6",
    "prod_max":       f"'{_ROI_SHEET}'!C7",
    "tempo_ei":       f"'{_ROI_SHEET}'!C8",
    "total_ei":       f"'{_ROI_SHEET}'!C9",
    "mensalidade":    f"'{_ROI_SHEET}'!C10",
    "meses":          f"'{_ROI_SHEET}'!C11",
    "saving_total":   f"'{_ROI_SHEET}'!C14",
    "saving_mensal":  f"'{_ROI_SHEET}'!C15",
    "roi_periodo":    f"'{_ROI_SHEET}'!C16",
    "roi_mensal":     f"'{_ROI_SHEET}'!C17",
    "custo_ei_digai": f"'{_ROI_SHEET}'!C18",
    "custo_ei_ta":    f"'{_ROI_SHEET}'!C19",
    "economia_ei":    f"'{_ROI_SHEET}'!C20",
}


# ─── Helper: conversão de índice 1-based para letra de coluna ─────────────────

def _col_letter(n: int) -> str:
    """Converte índice 1-based para letra(s) de coluna Excel (A, B, …, AA, …)."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


# ─── Helper: referência de coluna em 'Base de Dados' ─────────────────────────

def BD(col_src: str, col_map: dict) -> str:
    """Retorna referência de coluna inteira em 'Base de Dados', ex: '$A:$A'"""
    entry = col_map.get(col_src, {})
    letter = entry.get("letter", "A")
    return f"'{_BASE_SHEET}'!${letter}:${letter}"


# ─── Colunas canônicas ────────────────────────────────────────────────────────

_CANONICAL = [
    # ── Identificação ────────────────────────────────────────────────────────
    ("candidato_id",      "ID Candidato"),
    ("nome",              "Nome"),
    ("email",             "Email"),
    ("phone",             "Telefone"),
    ("cpf",               "CPF"),
    # ── Processo Seletivo (Gupy) ─────────────────────────────────────────────
    ("vaga",              "Vaga"),
    ("processo_seletivo", "Processo Seletivo"),   # Com DigAI / Sem DigAI
    ("status",            "Status"),              # Contratado / Reprovado / etc.
    ("data_cadastro",     "Data Cadastro"),
    ("data_final",        "Data Final"),
    ("data_contratacao",  "Data Contratacao"),
    ("_sla_dias",         "SLA Dias"),
    # ── DigAI Enriquecimento ─────────────────────────────────────────────────
    # Spec LOGICA_CRUZAMENTO PASSO 9: coluna H = "Passou pela DigAI" (Sim/Não)
    # Todas as fórmulas das outras abas filtram por esta coluna.
    ("_digai_realizado",  "Passou pela DigAI"),   # H — Sim/Não (era coluna M no gerar_relatorio.py)
    ("data_ei",           "Data Entrevista DigAI"),
    ("score_ia",          "Score DigAI"),
    ("aprovado_ia",       "Aprovado DigAI"),
    ("ranking_ia",        "Ranking DigAI"),
]


def _compute_col_map(df: pd.DataFrame) -> dict:
    """
    Constrói o mapeamento col_src -> {'idx', 'letter', 'label'}
    ANTES de criar o workbook (necessário para referências de fórmulas).
    """
    canonical = list(_CANONICAL)
    stage_cols = df.attrs.get("stage_cols", {})
    for n in sorted(stage_cols.keys()):
        for key, label_suffix in [
            ("name", "Nome"), ("entry", "Entrada"), ("exit", "Saida"), ("days", "Dias")
        ]:
            col_src = f"stage_{n}_{key}"
            if col_src in df.columns:
                canonical.append((col_src, f"Etapa {n} {label_suffix}"))

    cols = [
        (src, lbl) for src, lbl in canonical
        if src in df.columns or src == "_sla_dias"
    ]

    col_map = {}
    for idx_0based, (src, lbl) in enumerate(cols):
        col_map[src] = {
            "idx":    idx_0based,
            "letter": _col_letter(idx_0based + 1),
            "label":  lbl,
        }
    col_map["_cols_ordered"] = cols
    return col_map


# ─── Stubs de compatibilidade (excel_segmented.py usa openpyxl) ──────────────

def _fill(hex_color: str):
    """Stub: retorna dict bg_color para openpyxl via add_format."""
    return {"bg_color": hex_color}


def _font(bold=False, size=11, color="000000", italic=False):
    """Stub: retorna dict de fonte para openpyxl."""
    return {
        "font_name":  FONT_DEFAULT,
        "bold":       bold,
        "font_size":  size,
        "font_color": f"#{color}" if not color.startswith("#") else color,
        "italic":     italic,
    }


def _align(h="left", v="center", wrap=False):
    """Stub: retorna dict de alinhamento."""
    valign_map = {"center": "vcenter", "top": "top", "bottom": "bottom"}
    return {"align": h, "valign": valign_map.get(v, "vcenter"), "text_wrap": wrap}


def _border_thin():
    """Stub: retorna dict de borda thin."""
    return {"border": 1, "border_color": C_BORDER_STD}


# ─── Helpers de formato xlsxwriter ───────────────────────────────────────────

def _f(wb: xlsxwriter.Workbook, **props) -> object:
    """Cria format xlsxwriter a partir de kwargs. font_name padrão = Calibri."""
    props.setdefault("font_name", FONT_DEFAULT)
    return wb.add_format(props)


def _brd(props: dict) -> dict:
    """Adiciona borda padrão thin #9DC3E6 ao dict de props."""
    return {**props, "border": 1, "border_color": C_BORDER_STD}


def _brd_thick(props: dict) -> dict:
    """Adiciona borda medium #1F3864 ao dict de props."""
    return {**props, "border": 2, "border_color": C_BORDER_THICK}


def _brd_input(props: dict) -> dict:
    """Adiciona borda thin #FFD966 (células editáveis) ao dict de props."""
    return {**props, "border": 1, "border_color": C_AMARELO_BORDA}


# ─── Pré-processamento do DataFrame ──────────────────────────────────────────

def _preprocess_df(df: pd.DataFrame) -> pd.DataFrame:
    """Reset de índice e computa _sla_dias."""
    df = df.reset_index(drop=True)
    if "data_final" in df.columns and "data_cadastro" in df.columns:
        df["_sla_dias"] = (
            pd.to_datetime(df["data_final"], errors="coerce") -
            pd.to_datetime(df["data_cadastro"], errors="coerce")
        ).dt.days
    return df


# ─── Computação de dados para abas dinâmicas ─────────────────────────────────

def _compute_evolucao_mensal(df: pd.DataFrame) -> list:
    """Contratações mensais: total, Com DigAI, Sem DigAI, % via DigAI."""
    date_col = next(
        (c for c in ("data_contratacao", "data_final", "data_cadastro") if c in df.columns),
        None,
    )
    if date_col is None or "status" not in df.columns:
        return []

    contratados = df[df["status"].astype(str).str.contains("Contratad", na=False)].copy()
    if contratados.empty:
        return []

    contratados["_mes"] = (
        pd.to_datetime(contratados[date_col], errors="coerce")
        .dt.to_period("M")
    )
    contratados = contratados.dropna(subset=["_mes"])
    if contratados.empty:
        return []

    rows = []
    for mes, grp in contratados.groupby("_mes"):
        total = len(grp)
        com = 0
        if "processo_seletivo" in grp.columns:
            com = int(grp["processo_seletivo"].astype(str).str.contains("Com DigAI", na=False).sum())
        sem = total - com
        pct = com / total if total > 0 else 0.0
        rows.append({"mes": str(mes), "total": total, "com_digai": com,
                     "sem_digai": sem, "pct_digai": pct})

    return sorted(rows, key=lambda r: r["mes"])


def _compute_por_departamento(df: pd.DataFrame):
    """Contratações por área/departamento. Retorna None se coluna indisponível."""
    dept_col = next(
        (c for c in ("area", "Area", "Área", "departamento", "Departamento",
                     "filial", "Filial", "Unidade")
         if c in df.columns),
        None,
    )
    if dept_col is None or "status" not in df.columns:
        return None

    contratados = df[df["status"].astype(str).str.contains("Contratad", na=False)].copy()
    if contratados.empty:
        return None

    rows = []
    for dept, grp in contratados.groupby(dept_col):
        total = len(grp)
        via = 0
        if "processo_seletivo" in grp.columns:
            via = int(grp["processo_seletivo"].astype(str).str.contains("Com DigAI", na=False).sum())
        pct = via / total if total > 0 else 0.0
        rows.append({"departamento": str(dept), "total": total,
                     "via_digai": via, "pct_digai": pct})

    return sorted(rows, key=lambda r: -r["total"]) if rows else None


# ─── Computação de dados de segmentação ───────────────────────────────────────

def _compute_segmentacao_data(df: pd.DataFrame, dim_col: str) -> dict:
    """
    Computa KPIs por valor da dimensão e evolução mensal (Com DigAI global).

    Returns
    -------
    dict:
      kpi_rows   — lista de dicts por segmento
      period_rows— lista de dicts mensais globais
    """
    _COM  = "Com DigAI"
    _CONT = "Contratado"

    if dim_col not in df.columns:
        return {"kpi_rows": [], "period_rows": []}

    seg_values = sorted(
        df[dim_col].dropna().astype(str).str.strip().unique().tolist()
    )
    has_ps     = "processo_seletivo" in df.columns
    has_status = "status" in df.columns
    has_sla    = "_sla_dias" in df.columns

    kpi_rows = []
    for seg in seg_values:
        df_seg = df[df[dim_col].astype(str).str.strip() == seg]
        total  = len(df_seg)
        if total == 0:
            continue

        df_com    = df_seg[df_seg["processo_seletivo"] == _COM] if has_ps else df_seg.copy()
        total_com = len(df_com)
        adesao    = total_com / total if total > 0 else 0.0

        contratados   = 0
        assertividade = None
        if has_status and total_com > 0:
            contratados   = int((df_com["status"] == _CONT).sum())
            assertividade = contratados / total_com

        sla_medio = None
        if has_sla and has_status and total_com > 0:
            sla_s = df_com[df_com["status"] == _CONT]["_sla_dias"].dropna()
            if len(sla_s) > 0:
                sla_medio = round(float(sla_s.mean()), 1)

        kpi_rows.append({
            "segmento":      seg,
            "total":         total,
            "total_com":     total_com,
            "adesao":        adesao,
            "contratados":   contratados,
            "assertividade": assertividade,
            "volume_pct":    adesao,
            "sla_medio":     sla_medio,
        })

    # ── Evolução mensal (Com DigAI global) ──────────────────────────────
    period_rows = []
    date_col = next(
        (c for c in ("data_contratacao", "data_final", "data_cadastro") if c in df.columns),
        None,
    )
    if date_col and has_ps and has_status:
        df_com_all = df[df["processo_seletivo"] == _COM].copy()
        if not df_com_all.empty:
            df_com_all["_mes"] = (
                pd.to_datetime(df_com_all[date_col], errors="coerce")
                .dt.to_period("M")
                .astype(str)
            )
            for mes, grp in df_com_all.groupby("_mes"):
                if str(mes) == "NaT":
                    continue
                cont_mes = int((grp["status"] == _CONT).sum())
                period_rows.append({
                    "periodo":     str(mes),
                    "total_com":   len(grp),
                    "contratados": cont_mes,
                })
            period_rows.sort(key=lambda r: r["periodo"])

    return {"kpi_rows": kpi_rows, "period_rows": period_rows}


def _build_segmentacao_tab(
    wb: xlsxwriter.Workbook, ws, seg_data: dict, dim_label: str
) -> None:
    """
    Constrói aba de segmentação (xlsxwriter constant_memory compatível).

    Seção 1 — KPIs por segmento (destaque verde = melhor, vermelho = pior assertividade)
    Seção 2 — Performance por Período (Com DigAI global, mensal)
    """
    ws.hide_gridlines(2)
    ws.set_column(0, 0, 3)
    ws.set_column(1, 1, 32)
    ws.set_column(2, 2, 14)
    ws.set_column(3, 3, 13)
    ws.set_column(4, 4, 16)
    ws.set_column(5, 5, 16)
    ws.set_column(6, 6, 14)
    ws.set_column(7, 7, 13)
    ws.set_column(8, 8, 3)

    # ── Formatos ───────────────────────────────────────────────────────────
    _title = _f(wb, **_brd_thick({
        "bold": True, "font_size": 14,
        "font_color": C_BRANCO, "bg_color": C_AZUL_ESCURO,
        "align": "left", "valign": "vcenter",
    }))
    _hdr = _f(wb, **_brd({
        "bold": True, "font_size": 9,
        "font_color": C_BRANCO, "bg_color": C_AZUL_MEDIO,
        "align": "center", "valign": "vcenter",
    }))
    _section = _f(wb, **_brd({
        "bold": True, "font_size": 11,
        "font_color": C_BRANCO, "bg_color": C_AZUL_MEDIO,
        "align": "left", "valign": "vcenter",
    }))
    _seg_e = _f(wb, **_brd({
        "bold": True, "font_size": 9,
        "font_color": C_AZUL_ESCURO, "bg_color": C_AZUL_CLARO,
        "align": "left", "valign": "vcenter",
    }))
    _seg_o = _f(wb, **_brd({
        "bold": True, "font_size": 9,
        "font_color": C_AZUL_ESCURO, "bg_color": "#D6E8F7",
        "align": "left", "valign": "vcenter",
    }))
    _num_e = _f(wb, **_brd({
        "font_size": 9, "bg_color": C_BRANCO,
        "align": "center", "valign": "vcenter", "num_format": "#,##0",
    }))
    _num_o = _f(wb, **_brd({
        "font_size": 9, "bg_color": C_CINZA_CLARO,
        "align": "center", "valign": "vcenter", "num_format": "#,##0",
    }))
    _pct_e = _f(wb, **_brd({
        "font_size": 9, "bg_color": C_BRANCO,
        "align": "center", "valign": "vcenter", "num_format": "0.0%",
    }))
    _pct_o = _f(wb, **_brd({
        "font_size": 9, "bg_color": C_CINZA_CLARO,
        "align": "center", "valign": "vcenter", "num_format": "0.0%",
    }))
    _dec_e = _f(wb, **_brd({
        "font_size": 9, "bg_color": C_BRANCO,
        "align": "center", "valign": "vcenter", "num_format": "#,##0.0",
    }))
    _dec_o = _f(wb, **_brd({
        "font_size": 9, "bg_color": C_CINZA_CLARO,
        "align": "center", "valign": "vcenter", "num_format": "#,##0.0",
    }))
    _na_e = _f(wb, **_brd({
        "font_size": 9, "bg_color": C_CINZA_NA,
        "font_color": "#888888", "align": "center", "valign": "vcenter",
    }))
    _na_o = _f(wb, **_brd({
        "font_size": 9, "bg_color": "#CCCCCC",
        "font_color": "#888888", "align": "center", "valign": "vcenter",
    }))
    _best = _f(wb, **_brd({
        "bold": True, "font_size": 9,
        "bg_color": C_VERDE_CLARO, "font_color": C_VERDE_ESCURO,
        "align": "center", "valign": "vcenter", "num_format": "0.0%",
    }))
    _worst = _f(wb, **_brd({
        "bold": True, "font_size": 9,
        "bg_color": C_VERMELHO_CLARO, "font_color": C_VERMELHO,
        "align": "center", "valign": "vcenter", "num_format": "0.0%",
    }))

    row = 0
    kpi_rows    = seg_data.get("kpi_rows",    [])
    period_rows = seg_data.get("period_rows", [])

    # ── Espaçador ──────────────────────────────────────────────────────────
    ws.set_row(row, 8); row += 1

    # ── Título ─────────────────────────────────────────────────────────────
    ws.set_row(row, 36)
    ws.merge_range(row, 1, row, 7, f"Segmentação por {dim_label}", _title)
    row += 1

    # ── Espaçador ──────────────────────────────────────────────────────────
    ws.set_row(row, 8); row += 1

    # ─────────────────────────────────────────────────────────────────────
    # SEÇÃO 1: KPIs por segmento
    # ─────────────────────────────────────────────────────────────────────
    if kpi_rows:
        # Identifica melhor/pior assertividade para destaque
        valid_a = [
            (i, r["assertividade"])
            for i, r in enumerate(kpi_rows)
            if r.get("assertividade") is not None
        ]
        best_idx  = max(valid_a, key=lambda t: t[1])[0] if len(valid_a) > 1 else None
        worst_idx = min(valid_a, key=lambda t: t[1])[0] if len(valid_a) > 1 else None

        # Cabeçalhos
        ws.set_row(row, 30)
        for j, h in enumerate([
            "Segmento", "Total Cand.", "Adesão %",
            "Assertividade %", "Contratações DigAI",
            "Volume DigAI %", "SLA Médio (dias)",
        ]):
            ws.write(row, j + 1, h, _hdr)
        row += 1

        # Linhas de dados
        for i, rd in enumerate(kpi_rows):
            alt  = (i % 2 == 1)
            fn   = _num_o  if alt else _num_e
            fp   = _pct_o  if alt else _pct_e
            fd   = _dec_o  if alt else _dec_e
            fna  = _na_o   if alt else _na_e
            fseg = _seg_o  if alt else _seg_e

            ws.set_row(row, 20)
            ws.write(row, 1, rd["segmento"], fseg)
            ws.write(row, 2, rd["total"],    fn)

            a = rd.get("adesao")
            ws.write(row, 3, a if a is not None else "N/A",
                     fp if a is not None else fna)

            ass = rd.get("assertividade")
            if ass is not None:
                fa = _best if i == best_idx else (_worst if i == worst_idx else fp)
                ws.write(row, 4, ass, fa)
            else:
                ws.write(row, 4, "N/A", fna)

            ws.write(row, 5, rd.get("contratados", 0), fn)

            vol = rd.get("volume_pct")
            ws.write(row, 6, vol if vol is not None else "N/A",
                     fp if vol is not None else fna)

            sla = rd.get("sla_medio")
            ws.write(row, 7, sla if sla is not None else "N/A",
                     fd if sla is not None else fna)

            row += 1
    else:
        ws.set_row(row, 20)
        ws.write(row, 1, "Sem dados para segmentação.", _na_e)
        row += 1

    # ─────────────────────────────────────────────────────────────────────
    # SEÇÃO 2: Performance por Período (Com DigAI)
    # ─────────────────────────────────────────────────────────────────────
    if period_rows:
        ws.set_row(row, 14); row += 1

        ws.set_row(row, 30)
        ws.merge_range(row, 1, row, 4,
                       "Performance por Período (Com DigAI)", _section)
        row += 1

        ws.set_row(row, 26)
        for j, h in enumerate([
            "Período", "Total Com DigAI",
            "Contratações DigAI", "Taxa de Contratação",
        ]):
            ws.write(row, j + 1, h, _hdr)
        row += 1

        for i, pd_ in enumerate(period_rows):
            alt = (i % 2 == 1)
            fn  = _num_o if alt else _num_e
            fp  = _pct_o if alt else _pct_e
            fna = _na_o  if alt else _na_e

            ws.set_row(row, 18)
            ws.write(row, 1, pd_["periodo"],     fn)
            ws.write(row, 2, pd_["total_com"],   fn)
            ws.write(row, 3, pd_["contratados"], fn)

            tc   = pd_["total_com"]
            taxa = pd_["contratados"] / tc if tc > 0 else None
            ws.write(row, 4, taxa if taxa is not None else "N/A",
                     fp if taxa is not None else fna)
            row += 1


# ─── Aba 1: Indicadores DigAI ─────────────────────────────────────────────────

def _build_indicadores(
    wb: xlsxwriter.Workbook, ws,
    col_map: dict, relatorio: dict, params: dict, df: pd.DataFrame,
):
    """
    Aba 1 — KPIs, Saving, ROI, Definições.
    Layout (1-indexed Excel rows): col A=3, B=28, C=20, D=20, E=14, F=34, G=3
    """
    ws.hide_gridlines(2)
    ws.set_column(0, 0, 3)
    ws.set_column(1, 1, 28)
    ws.set_column(2, 2, 20)
    ws.set_column(3, 3, 20)
    ws.set_column(4, 4, 14)
    ws.set_column(5, 5, 34)
    ws.set_column(6, 6, 3)

    roi   = relatorio["roi"]
    kpis  = relatorio["kpis"]
    meta  = relatorio["meta"]
    n_ei  = roi.get("total_entrevistas_ia", 0)

    mensalidade = params.get("mensalidade_digai", 7600.0)
    n_meses     = roi.get("n_meses", 1)
    prod_max    = params.get("max_entrevistas_ta", 127)
    salario     = params.get("salario_ta_clt", 4750.0)

    # ── Formatos ─────────────────────────────────────────────────────────────
    f_title = _f(wb, **_brd_thick({
        "bold": True, "font_size": 14, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "center", "valign": "vcenter",
    }))
    f_saving_lbl = _f(wb, **_brd_thick({
        "bold": True, "font_size": 11, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "left", "valign": "vcenter",
    }))
    f_saving_val = _f(wb, **_brd_thick({
        "bold": True, "font_size": 14, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "center", "valign": "vcenter",
        "num_format": "R$ #,##0.00",
    }))
    f_roi_val = _f(wb, **_brd_thick({
        "bold": True, "font_size": 14, "font_color": C_BRANCO,
        "bg_color": C_VERDE_ESCURO, "align": "center", "valign": "vcenter",
        "num_format": '0.0"x"',
    }))
    f_note_blk = _f(wb, **_brd_thick({
        "font_size": 9, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "left", "valign": "vcenter",
        "text_wrap": True,
    }))
    f_kpi_hdr = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": C_BRANCO,
        "bg_color": C_AZUL_MEDIO, "align": "center", "valign": "vcenter",
        "text_wrap": True,
    }))
    f_kpi_lbl_w = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": "#000000",
        "bg_color": C_BRANCO, "align": "left", "valign": "vcenter",
    }))
    f_kpi_lbl_z = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": "#000000",
        "bg_color": C_CINZA_CLARO, "align": "left", "valign": "vcenter",
    }))
    f_com = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": C_AZUL_ESCURO,
        "bg_color": C_AZUL_CLARO, "align": "center", "valign": "vcenter",
    }))
    f_com_pct = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": C_AZUL_ESCURO,
        "bg_color": C_AZUL_CLARO, "align": "center", "valign": "vcenter",
        "num_format": "0.0%",
    }))
    f_com_curr = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": C_AZUL_ESCURO,
        "bg_color": C_AZUL_CLARO, "align": "center", "valign": "vcenter",
        "num_format": "R$ #,##0.00",
    }))
    f_com_dias = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": C_AZUL_ESCURO,
        "bg_color": C_AZUL_CLARO, "align": "center", "valign": "vcenter",
        "num_format": '0.0" dias"',
    }))
    f_ref = _f(wb, **_brd({
        "font_size": 10, "font_color": "#5C4800",
        "bg_color": C_AMARELO, "align": "center", "valign": "vcenter",
    }))
    f_ref_pct = _f(wb, **_brd({
        "font_size": 10, "font_color": "#5C4800",
        "bg_color": C_AMARELO, "align": "center", "valign": "vcenter",
        "num_format": "0.0%",
    }))
    f_ref_curr = _f(wb, **_brd({
        "font_size": 10, "font_color": "#5C4800",
        "bg_color": C_AMARELO, "align": "center", "valign": "vcenter",
        "num_format": "R$ #,##0.00",
    }))
    f_ref_dias = _f(wb, **_brd({
        "font_size": 10, "font_color": "#5C4800",
        "bg_color": C_AMARELO, "align": "center", "valign": "vcenter",
        "num_format": '0.0" dias"',
    }))
    f_na = _f(wb, **_brd({
        "italic": True, "font_size": 10, "font_color": "#595959",
        "bg_color": C_CINZA_NA, "align": "center", "valign": "vcenter",
    }))
    f_delta_neu = _f(wb, **_brd({
        "font_size": 10, "bg_color": C_CINZA_CLARO,
        "align": "center", "valign": "vcenter",
    }))
    f_def = _f(wb, **_brd({
        "italic": True, "font_size": 9, "font_color": "#595959",
        "bg_color": C_BRANCO, "align": "left", "valign": "vcenter",
        "text_wrap": True,
    }))
    f_def_z = _f(wb, **_brd({
        "italic": True, "font_size": 9, "font_color": "#595959",
        "bg_color": C_CINZA_CLARO, "align": "left", "valign": "vcenter",
        "text_wrap": True,
    }))
    f_defs_hdr = _f(wb, **_brd_thick({
        "bold": True, "font_size": 11, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "left", "valign": "vcenter",
        "indent": 2,
    }))
    f_def_body = _f(wb, **_brd({
        "font_size": 9, "font_color": "#3C3C3C",
        "bg_color": C_BRANCO, "align": "left", "valign": "vcenter",
        "text_wrap": True, "indent": 2,
    }))
    f_def_body_z = _f(wb, **_brd({
        "font_size": 9, "font_color": "#3C3C3C",
        "bg_color": C_CINZA_CLARO, "align": "left", "valign": "vcenter",
        "text_wrap": True, "indent": 2,
    }))
    f_delta_pos = wb.add_format(_brd({
        "font_name": FONT_DEFAULT, "bold": True, "font_size": 10,
        "font_color": C_VERDE_ESCURO, "bg_color": C_VERDE_CLARO,
        "align": "center", "valign": "vcenter", "num_format": "+0.0;-0.0;0",
    }))
    f_delta_neg = wb.add_format(_brd({
        "font_name": FONT_DEFAULT, "font_size": 10,
        "font_color": C_VERMELHO, "bg_color": C_VERMELHO_CLARO,
        "align": "center", "valign": "vcenter", "num_format": "+0.0;-0.0;0",
    }))

    # ── Row 0: spacer ──────────────────────────────────────────────────────
    ws.set_row(0, 8)

    # ── Row 1: Título (B:F merged, height 44) ─────────────────────────────
    ws.set_row(1, 44)
    title_text = (
        f"DigAI & {meta['cliente']}  |  Indicadores de Performance  |  "
        f"{meta['periodo']}"
    )
    ws.merge_range(1, 1, 1, 5, title_text, f_title)
    ws.write(1, 0, "", _f(wb, bg_color=C_AZUL_ESCURO))
    ws.write(1, 6, "", _f(wb, bg_color=C_AZUL_ESCURO))

    # ── Row 2: spacer ──────────────────────────────────────────────────────
    ws.set_row(2, 8)

    # ── Row 3: Saving block (height 30) ────────────────────────────────────
    ws.set_row(3, 30)
    ws.write(3, 1, "Saving DigAI", f_saving_lbl)
    ws.merge_range(3, 2, 3, 3,
                   f"={_ROI['saving_total']}", f_saving_val)
    note_saving = (
        f"({n_ei} entrevistas \u00f7 {prod_max} prod.m\u00e1x "
        f"\u00d7 R$ {salario:,.0f})"
    )
    ws.merge_range(3, 4, 3, 5, note_saving, f_note_blk)
    ws.write(3, 0, "", _f(wb, bg_color=C_AZUL_ESCURO))
    ws.write(3, 6, "", _f(wb, bg_color=C_AZUL_ESCURO))

    # ── Row 4: ROI block (height 30) ───────────────────────────────────────
    ws.set_row(4, 30)
    ws.write(4, 1, "ROI \u2014 DigAI", f_saving_lbl)
    ws.merge_range(4, 2, 4, 3,
                   f"={_ROI['roi_periodo']}", f_roi_val)
    note_roi = (
        f"Saving \u00f7 (R$ {mensalidade:,.0f} \u00d7 {n_meses} meses)"
    )
    ws.merge_range(4, 4, 4, 5, note_roi, f_note_blk)
    ws.write(4, 0, "", _f(wb, bg_color=C_AZUL_ESCURO))
    ws.write(4, 6, "", _f(wb, bg_color=C_AZUL_ESCURO))

    # ── Rows 5–6: spacers ──────────────────────────────────────────────────
    ws.set_row(5, 8)
    ws.set_row(6, 8)

    # ── Row 7: Cabeçalho tabela KPIs (height 24) ──────────────────────────
    ws.set_row(7, 24)
    for j, h in enumerate(["KPI", "Com DigAI", "Referencia", "\u0394", "Definicao"], start=1):
        ws.write(7, j, h, f_kpi_hdr)

    # ── KPI rows 8–15 ──────────────────────────────────────────────────────
    com_kpis = kpis.get("Com DigAI", {})
    sem_kpis = kpis.get("Sem DigAI", {})

    # Pré-computa vagas com DigAI
    vagas_com = 0
    vagas_total = 0
    if "vaga" in df.columns and "processo_seletivo" in df.columns:
        vagas_com   = df[df["processo_seletivo"].astype(str).str.contains("Com DigAI", na=False)]["vaga"].nunique()
        vagas_total = df["vaga"].nunique()

    has_score = (
        "score_ia" in df.columns and
        df["score_ia"].notna().any() and
        col_map.get("score_ia") is not None
    )
    has_sla = col_map.get("_sla_dias") is not None

    # (label, val_com, val_ref, val_delta, definicao, fmt_com, fmt_ref, fmt_delta)
    # fmt: "n"=número, "%"=percent, "R$"=currency, "d"=dias, None=str
    kpi_data = [
        (
            "Total entrevistas DigAI",
            n_ei,
            "N/A",
            None,
            "Quantidade de candidatos entrevistados pela IA DigAI no periodo",
            "n", None, None,
        ),
        (
            "Vagas com DigAI",
            vagas_com,
            vagas_total,
            None,
            "Vagas distintas que utilizaram DigAI no processo seletivo",
            "n", "n", None,
        ),
        (
            "Adesao (aprovacao IA)",
            (
                f'=IFERROR(COUNTIFS({BD("aprovado_ia", col_map)},"Sim",'
                f'{BD("_digai_realizado", col_map)},"Sim")/{n_ei},"N/A")'
                if n_ei > 0 and col_map.get("aprovado_ia") else
                (round(com_kpis.get("adesao", 0) / 100.0, 4)
                 if com_kpis.get("adesao") else "N/A")
            ),
            "N/A",
            None,
            "Candidatos que nao desistiram nem foram reprovados / total na EI",
            "%", None, None,
        ),
        (
            "Assertividade (contratacao)",
            (
                f'=IFERROR(COUNTIFS({BD("status", col_map)},"Contratado*",'
                f'{BD("_digai_realizado", col_map)},"Sim")/{n_ei},"N/A")'
                if n_ei > 0 else "N/A"
            ),
            "N/A",
            None,
            "% de aprovados que avancaram para contratacao",
            "%", None, None,
        ),
        (
            "Score medio \u2014 contratados",
            (
                f'=IFERROR(AVERAGEIFS({BD("score_ia", col_map)},'
                f'{BD("status", col_map)},"Contratado*",'
                f'{BD("_digai_realizado", col_map)},"Sim"),"N/A")'
                if has_score else "N/A"
            ),
            "N/A",
            None,
            ("Score medio dos candidatos contratados via DigAI" if has_score
             else "N/A — coluna score nao disponivel no arquivo DigAI"),
            "n", None, None,
        ),
        (
            "Contratados COM DigAI",
            (
                f'=COUNTIFS({BD("status", col_map)},"Contratado*",'
                f'{BD("_digai_realizado", col_map)},"Sim")'
            ),
            (
                f'=COUNTIFS({BD("status", col_map)},"Contratado*",'
                f'{BD("_digai_realizado", col_map)},"N\u00e3o")'
            ),
            "formula",
            "Total de contratacoes via processo Com DigAI vs Sem DigAI",
            "n", "n", "n",
        ),
        (
            "SLA medio \u2014 COM DigAI",
            (
                f'=IFERROR(AVERAGEIFS({BD("_sla_dias", col_map)},'
                f'{BD("status", col_map)},"Contratado*",'
                f'{BD("_digai_realizado", col_map)},"Sim"),"N/A")'
                if has_sla else "N/A"
            ),
            (
                f'=IFERROR(AVERAGEIFS({BD("_sla_dias", col_map)},'
                f'{BD("status", col_map)},"Contratado*",'
                f'{BD("_digai_realizado", col_map)},"N\u00e3o"),"N/A")'
                if has_sla else "N/A"
            ),
            ("formula_reversed" if has_sla else None),
            ("Media de dias da inscricao ate contratacao (menor = melhor)" if has_sla
             else "N/A — datas de SLA ausentes"),
            "d", "d", "d",
        ),
        (
            "Custo/EI DigAI vs TA",
            f"={_ROI['custo_ei_digai']}",
            f"={_ROI['custo_ei_ta']}",
            f"={_ROI['economia_ei']}",
            "Custo de uma entrevista via DigAI vs custo equivalente de recrutador TA",
            "R$", "R$", "R$",
        ),
    ]

    FIRST_KPI_ROW = 8
    for i, (label, val_com, val_ref, val_delta, defi, fc, fr, fd) in enumerate(kpi_data):
        r = FIRST_KPI_ROW + i
        ws.set_row(r, 22)
        alt = i % 2 == 1

        # Formato KPI label
        lbl_fmt = f_kpi_lbl_z if alt else f_kpi_lbl_w

        # Formatos Com DigAI
        fmt_com_map = {"%": f_com_pct, "R$": f_com_curr, "d": f_com_dias, "n": f_com}
        fmt_ref_map = {"%": f_ref_pct, "R$": f_ref_curr, "d": f_ref_dias, "n": f_ref}

        ws.write(r, 1, label, lbl_fmt)

        # Col C: Com DigAI
        if isinstance(val_com, str) and val_com.startswith("="):
            ws.write_formula(r, 2, val_com, fmt_com_map.get(fc, f_com))
        elif val_com == "N/A":
            ws.write(r, 2, "N/A", f_na)
        else:
            ws.write(r, 2, val_com, fmt_com_map.get(fc, f_com))

        # Col D: Referência
        if isinstance(val_ref, str) and val_ref.startswith("="):
            ws.write_formula(r, 3, val_ref, fmt_ref_map.get(fr, f_ref))
        elif val_ref == "N/A":
            ws.write(r, 3, "N/A", f_na)
        else:
            ws.write(r, 3, val_ref, fmt_ref_map.get(fr, f_ref))

        # Col E: Δ
        if val_delta is None:
            ws.write(r, 4, "\u2014", f_delta_neu)
        elif val_delta == "formula":
            # Contratados: delta = Com - Ref (positive = more with DigAI)
            ws.write_formula(r, 4, f"=C{r+1}-D{r+1}", f_delta_neu)
        elif val_delta == "formula_reversed":
            # SLA: delta = Ref - Com (positive = DigAI is faster = good)
            ws.write_formula(r, 4, f"=D{r+1}-C{r+1}", f_delta_neu)
        elif isinstance(val_delta, str) and val_delta.startswith("="):
            ws.write_formula(r, 4, val_delta, f_delta_neu)
        else:
            ws.write(r, 4, val_delta, f_delta_neu)

        # Col F: Definição
        def_fmt = f_def_z if alt else f_def
        ws.write(r, 5, defi, def_fmt)

    # Conditional formatting on col E (Δ column, index 4)
    last_kpi_row = FIRST_KPI_ROW + len(kpi_data) - 1
    ws.conditional_format(FIRST_KPI_ROW, 4, last_kpi_row, 4, {
        "type": "cell", "criteria": ">", "value": 0, "format": f_delta_pos,
    })
    ws.conditional_format(FIRST_KPI_ROW, 4, last_kpi_row, 4, {
        "type": "cell", "criteria": "<", "value": 0, "format": f_delta_neg,
    })

    # ── Espaçador antes definições ────────────────────────────────────────
    spacer_row = last_kpi_row + 1
    ws.set_row(spacer_row, 8)

    # ── Header Definições KPIs ────────────────────────────────────────────
    defs_hdr_row = spacer_row + 1
    ws.set_row(defs_hdr_row, 24)
    ws.merge_range(defs_hdr_row, 1, defs_hdr_row, 5, "Definicoes KPIs", f_defs_hdr)
    ws.write(defs_hdr_row, 0, "", _f(wb, bg_color=C_AZUL_ESCURO))
    ws.write(defs_hdr_row, 6, "", _f(wb, bg_color=C_AZUL_ESCURO))

    # ── Linhas de definições ──────────────────────────────────────────────
    definitions = [
        "Adesao: Candidatos que nao desistiram nem foram reprovados / total na Entrevista Inteligente",
        "Assertividade: % de aprovados que avancaram para Triagem ou etapas subsequentes",
        "SLA: Media de dias da jornada do candidato — da inscricao ate a contratacao",
        "Produtividade maxima por recrutador: ~127 entrevistas/mes (60% do tempo produtivo)",
        "Custo mensal de um recrutador: salario bruto R$ 4.750 + encargos (~CLT)",
        "Custo por entrevista: calculado automaticamente na aba Calculadora ROI",
        "Celulas em amarelo = valores de referencia historica (Sem DigAI) para comparacao",
        "Formulas usam wildcard para capturar todas as variacoes de grafia do ATS",
    ]
    for i, defi in enumerate(definitions):
        dr = defs_hdr_row + 1 + i
        ws.set_row(dr, 20)
        alt = i % 2 == 1
        df_fmt = f_def_body_z if alt else f_def_body
        ws.merge_range(dr, 1, dr, 5, defi, df_fmt)


# ─── Aba 2: Calculadora ROI ──────────────────────────────────────────────────

def _build_calculadora_roi(
    wb: xlsxwriter.Workbook, ws,
    col_map: dict, relatorio: dict, params: dict,
):
    """
    Aba 2 — parâmetros editáveis (amarelo) + resultados calculados.
    Células editáveis: C6, C7, C8, C10, C11 (1-indexed).
    Célula calculada da Base: C9 (Total EI DigAI via COUNTIF).
    Resultados: C14–C20.
    """
    ws.hide_gridlines(2)
    ws.set_column(0, 0, 3)
    ws.set_column(1, 1, 34)
    ws.set_column(2, 2, 22)
    ws.set_column(3, 3, 32)
    ws.set_column(4, 4, 3)

    roi    = relatorio["roi"]
    n_ei   = roi.get("total_entrevistas_ia", 0)
    n_mes  = roi.get("n_meses", 1)
    meta   = relatorio["meta"]

    mensalidade = params.get("mensalidade_digai", 7600.0)
    salario     = params.get("salario_ta_clt", 4750.0)
    prod_max    = params.get("max_entrevistas_ta", 127)
    tempo_ei    = params.get("tempo_entrevista_min", 30)

    # ── Formatos ─────────────────────────────────────────────────────────────
    f_title = _f(wb, **_brd_thick({
        "bold": True, "font_size": 14, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "center", "valign": "vcenter",
    }))
    f_sec = _f(wb, **_brd({
        "bold": True, "font_size": 11, "font_color": C_BRANCO,
        "bg_color": C_AZUL_MEDIO, "align": "left", "valign": "vcenter",
        "indent": 1,
    }))
    f_col_hdr = _f(wb, **_brd({
        "bold": True, "font_size": 9, "font_color": C_BRANCO,
        "bg_color": C_AZUL_MEDIO, "align": "center", "valign": "vcenter",
    }))
    f_lbl_w = _f(wb, **_brd({
        "font_size": 10, "font_color": "#000000",
        "bg_color": C_BRANCO, "align": "left", "valign": "vcenter",
    }))
    f_lbl_z = _f(wb, **_brd({
        "font_size": 10, "font_color": "#000000",
        "bg_color": C_CINZA_CLARO, "align": "left", "valign": "vcenter",
    }))
    f_note_w = _f(wb, **_brd({
        "italic": True, "font_size": 9, "font_color": "#595959",
        "bg_color": C_BRANCO, "align": "left", "valign": "vcenter", "text_wrap": True,
    }))
    f_note_z = _f(wb, **_brd({
        "italic": True, "font_size": 9, "font_color": "#595959",
        "bg_color": C_CINZA_CLARO, "align": "left", "valign": "vcenter", "text_wrap": True,
    }))
    # Células editáveis: amarelo
    def f_input(num_format=None):
        p = _brd_input({
            "font_name": FONT_DEFAULT, "bold": True, "font_size": 10,
            "font_color": "#000000", "bg_color": C_AMARELO,
            "align": "right", "valign": "vcenter",
        })
        if num_format:
            p["num_format"] = num_format
        return wb.add_format(p)

    # Célula calculada da Base: azul claro
    def f_calc(num_format=None):
        p = _brd({
            "font_name": FONT_DEFAULT, "bold": True, "font_size": 10,
            "font_color": C_AZUL_ESCURO, "bg_color": C_AZUL_CLARO,
            "align": "right", "valign": "vcenter",
        })
        if num_format:
            p["num_format"] = num_format
        return wb.add_format(p)

    def f_result(bg=C_AZUL_CLARO, bold=True, size=10, font_color=C_AZUL_ESCURO,
                 num_format=None):
        p = _brd({
            "font_name": FONT_DEFAULT, "bold": bold, "font_size": size,
            "font_color": font_color, "bg_color": bg,
            "align": "right", "valign": "vcenter",
        })
        if num_format:
            p["num_format"] = num_format
        return wb.add_format(p)

    # ── Row 0: spacer ──────────────────────────────────────────────────────
    ws.set_row(0, 8)

    # ── Row 1: Título ──────────────────────────────────────────────────────
    ws.set_row(1, 44)
    ws.merge_range(1, 1, 1, 3,
                   f"Calculadora ROI \u2014 DigAI  |  {meta['cliente']}  |  {meta['periodo']}",
                   f_title)

    # ── Row 2: spacer ──────────────────────────────────────────────────────
    ws.set_row(2, 8)

    # ── Row 3: Section PARÂMETROS ──────────────────────────────────────────
    ws.set_row(3, 20)
    ws.merge_range(3, 1, 3, 3, "PAR\u00c2METROS", f_sec)

    # ── Row 4: Col headers ─────────────────────────────────────────────────
    ws.set_row(4, 22)
    for j, h in enumerate(["Parametro", "Valor", "Nota"], start=1):
        ws.write(4, j, h, f_col_hdr)

    # ── Rows 5–10: Parâmetros (Excel rows 6–11, 0-indexed 5–10) ───────────
    parem_rows = [
        # (label, value, fmt, note, editable)
        ("Salario base recrutador \u2014 CLT", salario,   "R$ #,##0.00", "Editavel: salario bruto mensal do TA", True),
        ("Produtividade maxima (EI/mes)",       prod_max, "#,##0",       "22 dias x 8h x 60% produtividade / 30 min/EI", True),
        ("Tempo de entrevista (min)",           tempo_ei, "0",           "Duracao media de uma entrevista presencial (min)", True),
        ("Total entrevistas DigAI",             n_ei,     "#,##0",       f"COUNTIF direto na Base de Dados (nao editavel)", False),
        ("Mensalidade DigAI",                   mensalidade, "R$ #,##0.00", "Editavel: valor mensal pago ao DigAI (R$)", True),
        ("Meses no periodo",                    n_mes,    "0",           "Auto-detectado pelo periodo informado", True),
    ]
    for i, (label, value, num_fmt, note, editable) in enumerate(parem_rows):
        r = 5 + i  # 0-indexed row 5 = Excel row 6
        ws.set_row(r, 22)
        alt = i % 2 == 1
        lbl_fmt  = f_lbl_z if alt else f_lbl_w
        note_fmt = f_note_z if alt else f_note_w

        ws.write(r, 1, label, lbl_fmt)

        if editable:
            ws.write(r, 2, value, f_input(num_fmt))
        else:
            # Total EI DigAI: valor Python direto (COUNTIF cross-sheet
            # não é confiável com constant_memory — sem cached value)
            ws.write(r, 2, value, f_calc(num_fmt))

        ws.write(r, 3, note, note_fmt)

    # ── Row 11: spacer ─────────────────────────────────────────────────────
    ws.set_row(11, 8)

    # ── Row 12: Section RESULTADOS ─────────────────────────────────────────
    ws.set_row(12, 20)
    ws.merge_range(12, 1, 12, 3, "RESULTADOS CALCULADOS", f_sec)

    # ── Rows 13–19: Resultados (Excel 14–20, 0-indexed 13–19) ─────────────
    # Referências 1-indexed para as células de parâmetros acima:
    # C6=salario, C7=prodmax, C8=tempo_ei, C9=total_ei, C10=mensalidade, C11=meses
    result_rows = [
        ("Saving Total (periodo)",
         "=IFERROR((C9/C7)*C6-C10,\"N/A\")",
         "R$ #,##0.00",
         "(EI_TOTAL / PROD_MAX) x CLT - Mensalidade DigAI",
         C_AZUL_CLARO, C_AZUL_ESCURO, False),
        ("Saving Mensal medio",
         "=IFERROR(C14/C11,\"N/A\")",
         "R$ #,##0.00",
         "Saving Total / Meses no periodo",
         C_AZUL_CLARO, C_AZUL_ESCURO, False),
        ("ROI do periodo",
         "=IFERROR(C14/(C10*C11),\"N/A\")",
         '0.0"x"',
         "Saving / (Mensalidade x Meses)",
         C_VERDE_ESCURO, C_BRANCO, True),
        ("ROI mensal estimado",
         "=IFERROR((C14/C11)/C10,\"N/A\")",
         '0.0"x"',
         "(Saving / Meses) / Mensalidade",
         C_VERDE_ESCURO, C_BRANCO, True),
        ("Custo/EI pelo DigAI",
         "=IFERROR(C10/C9,\"N/A\")",
         "R$ #,##0.00",
         "Mensalidade / Total EI DigAI",
         C_AZUL_CLARO, C_AZUL_ESCURO, False),
        ("Custo/EI pelo TA",
         "=IFERROR(C6/C7,\"N/A\")",
         "R$ #,##0.00",
         "Salario CLT / Produtividade maxima",
         C_AMARELO, "#5C4800", False),
        ("Economia por entrevista",
         "=IFERROR(C19-C18,\"N/A\")",
         "R$ #,##0.00",
         "Custo TA - Custo DigAI",
         C_VERDE_CLARO, C_VERDE_ESCURO, True),
    ]
    for i, (label, formula, num_fmt, note, bg, fc, is_big) in enumerate(result_rows):
        r = 13 + i  # 0-indexed
        ws.set_row(r, 22)
        alt = i % 2 == 1
        lbl_fmt  = f_lbl_z if alt else f_lbl_w
        note_fmt = f_note_z if alt else f_note_w

        ws.write(r, 1, label, lbl_fmt)
        ws.write_formula(r, 2, formula, f_result(
            bg=bg, bold=is_big, size=12 if is_big else 10,
            font_color=fc, num_format=num_fmt,
        ))
        ws.write(r, 3, note, note_fmt)

    # ── Row 20: spacer ─────────────────────────────────────────────────────
    ws.set_row(20, 8)

    # ── Row 21: Nota rodapé ────────────────────────────────────────────────
    ws.set_row(21, 20)
    f_rodape = _f(wb, **_brd_input({
        "italic": True, "font_size": 9, "font_color": "#4A4A4A",
        "bg_color": C_AMARELO, "align": "left", "valign": "vcenter",
    }))
    ws.merge_range(21, 1, 21, 3,
                   "\U0001f4a1  Celulas em amarelo sao editaveis. "
                   "As demais sao calculadas automaticamente.",
                   f_rodape)


# ─── Aba 3: Score & Qualidade (condicional) ───────────────────────────────────

def _build_score_qualidade(wb: xlsxwriter.Workbook, ws, df: pd.DataFrame):
    """Aba 3 — distribuição de score IA (apenas quando score disponível)."""
    ws.hide_gridlines(2)
    ws.set_column(0, 0, 3)
    ws.set_column(1, 1, 28)
    ws.set_column(2, 2, 20)
    ws.set_column(3, 3, 20)
    ws.set_column(4, 4, 14)
    ws.set_column(5, 5, 3)

    f_title = _f(wb, **_brd_thick({
        "bold": True, "font_size": 13, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "center", "valign": "vcenter",
    }))
    f_sec = _f(wb, **_brd({
        "bold": True, "font_size": 11, "font_color": C_BRANCO,
        "bg_color": C_AZUL_MEDIO, "align": "left", "valign": "vcenter",
    }))
    f_hdr = _f(wb, **_brd({
        "bold": True, "font_size": 9, "font_color": C_BRANCO,
        "bg_color": C_AZUL_MEDIO, "align": "center", "valign": "vcenter",
    }))

    # Título
    ws.set_row(0, 8)
    ws.set_row(1, 36)
    ws.merge_range(1, 1, 1, 4, "Score & Qualidade — DigAI", f_title)

    # Bloco 1: comparativo de score
    ws.set_row(3, 20)
    ws.merge_range(3, 1, 3, 4, "Comparativo de Score", f_sec)
    ws.set_row(4, 22)
    for j, h in enumerate(["Metrica", "Contratados DigAI", "Todos Entrevistados", "\u0394"], start=1):
        ws.write(4, j, h, f_hdr)

    contratados_digai = df[
        (df["status"].astype(str).str.contains("Contratad", na=False)) &
        (df["processo_seletivo"].astype(str).str.contains("Com DigAI", na=False))
    ]["score_ia"].dropna()
    todos_ei = df["score_ia"].dropna()

    score_metrics = [
        ("Score medio",   contratados_digai.mean(),   todos_ei.mean()),
        ("Score mediana", contratados_digai.median(),  todos_ei.median()),
        ("Score minimo",  contratados_digai.min(),     todos_ei.min()),
        ("Score maximo",  contratados_digai.max(),     todos_ei.max()),
        ("Total",         len(contratados_digai),      len(todos_ei)),
    ]
    f_cd = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": C_AZUL_ESCURO,
        "bg_color": C_AZUL_CLARO, "align": "center", "valign": "vcenter",
        "num_format": "0.00",
    }))
    f_al = _f(wb, **_brd({
        "font_size": 10, "font_color": "#5C4800",
        "bg_color": C_AMARELO, "align": "center", "valign": "vcenter",
        "num_format": "0.00",
    }))
    f_lbl = _f(wb, **_brd({
        "font_size": 10, "bg_color": C_BRANCO,
        "align": "left", "valign": "vcenter",
    }))
    for i, (label, v_cont, v_all) in enumerate(score_metrics):
        r = 5 + i
        ws.set_row(r, 20)
        ws.write(r, 1, label, f_lbl)
        ws.write(r, 2, round(float(v_cont), 2) if not (isinstance(v_cont, float) and math.isnan(v_cont)) else "N/A", f_cd)
        ws.write(r, 3, round(float(v_all), 2) if not (isinstance(v_all, float) and math.isnan(v_all)) else "N/A", f_al)
        delta = (v_cont - v_all) if (
            isinstance(v_cont, (int, float)) and not math.isnan(v_cont) and
            isinstance(v_all, (int, float)) and not math.isnan(v_all)
        ) else None
        f_delta = _f(wb, **_brd({
            "font_size": 10, "align": "center", "valign": "vcenter",
            "bg_color": C_VERDE_CLARO if delta is not None and delta > 0 else
                        C_VERMELHO_CLARO if delta is not None and delta < 0 else
                        C_CINZA_CLARO,
            "font_color": C_VERDE_ESCURO if delta is not None and delta > 0 else
                          C_VERMELHO if delta is not None and delta < 0 else "#595959",
            "num_format": "+0.00;-0.00;0",
        }))
        ws.write(r, 4, round(float(delta), 2) if delta is not None else "\u2014", f_delta)

    # Bloco 2: distribuição por faixa
    faixa_row = 5 + len(score_metrics) + 2
    ws.set_row(faixa_row, 20)
    ws.merge_range(faixa_row, 1, faixa_row, 4, "Distribuicao por Faixa de Score", f_sec)

    faixas = [
        ("Alta  (8\u201310)",  8.0,  10.1, C_VERDE_CLARO,   C_VERDE_ESCURO),
        ("Media (6\u20138)",   6.0,   8.0, C_AZUL_CLARO,    C_AZUL_ESCURO),
        ("Baixa (0\u20136)",   0.0,   6.0, C_VERMELHO_CLARO, C_VERMELHO),
    ]
    for j, h in enumerate(["Faixa", "Contratados", "Todos Entrevistados", "%"], start=1):
        ws.write(faixa_row + 1, j, h, f_hdr)
    for i, (label, lo, hi, bg, fc) in enumerate(faixas):
        r = faixa_row + 2 + i
        ws.set_row(r, 20)
        mask_cont = (contratados_digai >= lo) & (contratados_digai < hi)
        mask_all  = (todos_ei >= lo) & (todos_ei < hi)
        n_cont = int(mask_cont.sum())
        n_all  = int(mask_all.sum())
        pct    = n_cont / len(contratados_digai) if len(contratados_digai) > 0 else 0

        f_faixa = _f(wb, **_brd({"font_size": 10, "bg_color": bg, "font_color": fc,
                                  "align": "left", "valign": "vcenter", "bold": True}))
        f_num   = _f(wb, **_brd({"font_size": 10, "bg_color": bg, "font_color": fc,
                                  "align": "center", "valign": "vcenter",
                                  "num_format": "#,##0"}))
        f_pct_f = _f(wb, **_brd({"font_size": 10, "bg_color": bg, "font_color": fc,
                                   "align": "center", "valign": "vcenter",
                                   "num_format": "0.0%"}))
        ws.write(r, 1, label, f_faixa)
        ws.write(r, 2, n_cont, f_num)
        ws.write(r, 3, n_all,  f_num)
        ws.write(r, 4, pct,    f_pct_f)


# ─── Aba 4: Evolução Mensal ───────────────────────────────────────────────────

def _build_evolucao_mensal_tab(wb: xlsxwriter.Workbook, ws, data: list):
    """Aba 4 — contratações mês a mês."""
    ws.hide_gridlines(2)
    ws.set_column(0, 0, 3)
    ws.set_column(1, 1, 16)
    ws.set_column(2, 2, 14)
    ws.set_column(3, 3, 14)
    ws.set_column(4, 4, 14)
    ws.set_column(5, 5, 14)
    ws.set_column(6, 6, 3)

    f_title = _f(wb, **_brd_thick({
        "bold": True, "font_size": 13, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "center", "valign": "vcenter",
    }))
    f_hdr = _f(wb, **_brd({
        "bold": True, "font_size": 9, "font_color": C_BRANCO,
        "bg_color": C_AZUL_MEDIO, "align": "center", "valign": "vcenter",
    }))
    f_mes_w = _f(wb, **_brd({
        "bold": True, "font_size": 10, "bg_color": C_BRANCO,
        "align": "center", "valign": "vcenter",
    }))
    f_mes_z = _f(wb, **_brd({
        "bold": True, "font_size": 10, "bg_color": C_CINZA_CLARO,
        "align": "center", "valign": "vcenter",
    }))
    f_tot_w = _f(wb, **_brd({
        "font_size": 10, "bg_color": C_BRANCO, "align": "center", "valign": "vcenter",
        "num_format": "#,##0",
    }))
    f_tot_z = _f(wb, **_brd({
        "font_size": 10, "bg_color": C_CINZA_CLARO, "align": "center", "valign": "vcenter",
        "num_format": "#,##0",
    }))
    f_com_v = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": C_AZUL_ESCURO,
        "bg_color": C_AZUL_CLARO, "align": "center", "valign": "vcenter",
        "num_format": "#,##0",
    }))
    f_sem_v = _f(wb, **_brd({
        "font_size": 10, "font_color": "#5C4800",
        "bg_color": C_AMARELO, "align": "center", "valign": "vcenter",
        "num_format": "#,##0",
    }))
    f_tot_row = _f(wb, **_brd_thick({
        "bold": True, "font_size": 10, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "center", "valign": "vcenter",
        "num_format": "#,##0",
    }))
    f_tot_pct = _f(wb, **_brd_thick({
        "bold": True, "font_size": 10, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "center", "valign": "vcenter",
        "num_format": "0.0%",
    }))

    ws.set_row(0, 8)
    ws.set_row(1, 34)
    ws.merge_range(1, 1, 1, 5, "Evolucao Mensal de Contratacoes", f_title)
    ws.set_row(2, 8)
    ws.set_row(3, 22)
    for j, h in enumerate(["Mes", "Total", "Com DigAI", "Sem DigAI", "% via DigAI"], start=1):
        ws.write(3, j, h, f_hdr)

    tot_t = tot_c = tot_s = 0
    for i, row_d in enumerate(data):
        r = 4 + i
        ws.set_row(r, 20)
        alt = i % 2 == 1
        pct = row_d["pct_digai"]

        ws.write(r, 1, row_d["mes"], f_mes_z if alt else f_mes_w)
        ws.write(r, 2, row_d["total"], f_tot_z if alt else f_tot_w)
        ws.write(r, 3, row_d["com_digai"], f_com_v)
        ws.write(r, 4, row_d["sem_digai"], f_sem_v)

        # % via DigAI: >= 15% → verde, < 15% → cinza
        if pct >= 0.15:
            f_pct = _f(wb, **_brd({
                "bold": True, "font_size": 10, "font_color": C_VERDE_ESCURO,
                "bg_color": C_VERDE_CLARO, "align": "center", "valign": "vcenter",
                "num_format": "0.0%",
            }))
        else:
            f_pct = _f(wb, **_brd({
                "font_size": 10, "font_color": "#595959",
                "bg_color": C_CINZA_CLARO, "align": "center", "valign": "vcenter",
                "num_format": "0.0%",
            }))
        ws.write(r, 5, pct, f_pct)
        tot_t += row_d["total"]
        tot_c += row_d["com_digai"]
        tot_s += row_d["sem_digai"]

    # Linha de totais
    tot_r = 4 + len(data)
    ws.set_row(tot_r, 22)
    ws.write(tot_r, 1, "TOTAL", f_tot_row)
    ws.write(tot_r, 2, tot_t, f_tot_row)
    ws.write(tot_r, 3, tot_c, f_tot_row)
    ws.write(tot_r, 4, tot_s, f_tot_row)
    ws.write(tot_r, 5, tot_c / tot_t if tot_t > 0 else 0, f_tot_pct)


# ─── Aba 5: Por Departamento (condicional) ────────────────────────────────────

def _build_por_departamento_tab(wb: xlsxwriter.Workbook, ws, data: list):
    """Aba 5 — contratações por área/departamento."""
    ws.hide_gridlines(2)
    ws.set_column(0, 0, 3)
    ws.set_column(1, 1, 28)
    ws.set_column(2, 2, 16)
    ws.set_column(3, 3, 16)
    ws.set_column(4, 4, 14)
    ws.set_column(5, 5, 3)

    f_title = _f(wb, **_brd_thick({
        "bold": True, "font_size": 13, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "center", "valign": "vcenter",
    }))
    f_hdr = _f(wb, **_brd({
        "bold": True, "font_size": 9, "font_color": C_BRANCO,
        "bg_color": C_AZUL_MEDIO, "align": "center", "valign": "vcenter",
    }))

    ws.set_row(0, 8)
    ws.set_row(1, 34)
    ws.merge_range(1, 1, 1, 4, "Contratacoes por Departamento / Area", f_title)
    ws.set_row(2, 8)
    ws.set_row(3, 22)
    for j, h in enumerate(["Departamento", "Total Contratados", "Via DigAI", "% via DigAI"], start=1):
        ws.write(3, j, h, f_hdr)

    f_dept_w = _f(wb, **_brd({"bold": True, "font_size": 10, "bg_color": C_BRANCO,
                               "align": "left", "valign": "vcenter"}))
    f_dept_z = _f(wb, **_brd({"bold": True, "font_size": 10, "bg_color": C_CINZA_CLARO,
                               "align": "left", "valign": "vcenter"}))
    f_tot_w = _f(wb, **_brd({"font_size": 10, "bg_color": C_BRANCO, "align": "center",
                              "valign": "vcenter", "num_format": "#,##0"}))
    f_tot_z = _f(wb, **_brd({"font_size": 10, "bg_color": C_CINZA_CLARO, "align": "center",
                              "valign": "vcenter", "num_format": "#,##0"}))

    for i, row_d in enumerate(data):
        r = 4 + i
        ws.set_row(r, 20)
        alt = i % 2 == 1
        via = row_d["via_digai"]
        pct = row_d["pct_digai"]

        ws.write(r, 1, row_d["departamento"], f_dept_z if alt else f_dept_w)
        ws.write(r, 2, row_d["total"], f_tot_z if alt else f_tot_w)

        if via > 0:
            ws.write(r, 3, via, _f(wb, **_brd({
                "bold": True, "font_size": 10, "font_color": C_AZUL_ESCURO,
                "bg_color": C_AZUL_CLARO, "align": "center", "valign": "vcenter",
                "num_format": "#,##0",
            })))
            ws.write(r, 4, pct, _f(wb, **_brd({
                "bold": True, "font_size": 10, "font_color": C_VERDE_ESCURO,
                "bg_color": C_VERDE_CLARO, "align": "center", "valign": "vcenter",
                "num_format": "0.0%",
            })))
        else:
            f_zero = _f(wb, **_brd({
                "font_size": 10, "font_color": "#595959",
                "bg_color": C_CINZA_CLARO if alt else C_BRANCO,
                "align": "center", "valign": "vcenter",
            }))
            ws.write(r, 3, via, f_zero)
            ws.write(r, 4, "\u2014", f_zero)


# ─── Aba 3 (nova): Por Time ──────────────────────────────────────────────────

def _build_por_time(
    wb: xlsxwriter.Workbook, ws,
    col_map: dict, relatorio: dict, df: pd.DataFrame,
) -> None:
    """
    Aba 3 — Comparativo Com DigAI vs Sem DigAI.
    Usa formulas COUNTIFS/AVERAGEIFS na Base de Dados.
    """
    ws.hide_gridlines(2)
    ws.set_column(0, 0, 3)
    ws.set_column(1, 1, 32)
    ws.set_column(2, 2, 18)
    ws.set_column(3, 3, 18)
    ws.set_column(4, 4, 16)
    ws.set_column(5, 5, 3)

    meta   = relatorio["meta"]
    cliente = meta.get("cliente", "")
    periodo = meta.get("periodo", "")

    has_sla   = col_map.get("_sla_dias") is not None
    has_score = col_map.get("score_ia") is not None

    # ── Formatos ─────────────────────────────────────────────────────────────
    f_title = _f(wb, **_brd_thick({
        "bold": True, "font_size": 14, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "left", "valign": "vcenter",
    }))
    f_hdr_kpi = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": C_BRANCO,
        "bg_color": C_AZUL_MEDIO, "align": "center", "valign": "vcenter",
    }))
    f_hdr_com = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": C_AZUL_ESCURO,
        "bg_color": C_AZUL_CLARO, "align": "center", "valign": "vcenter",
    }))
    f_hdr_sem = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": "#5C4800",
        "bg_color": C_AMARELO, "align": "center", "valign": "vcenter",
    }))
    f_hdr_dif = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": C_BRANCO,
        "bg_color": C_AZUL_MEDIO, "align": "center", "valign": "vcenter",
    }))

    def _lbl_fmt(alt):
        return _f(wb, **_brd({
            "bold": True, "font_size": 10, "font_color": "#000000",
            "bg_color": C_CINZA_CLARO if alt else C_BRANCO,
            "align": "left", "valign": "vcenter",
        }))

    def _com_fmt(alt, num_format=None):
        p = _brd({
            "bold": True, "font_size": 10, "font_color": C_AZUL_ESCURO,
            "bg_color": C_AZUL_CLARO, "align": "center", "valign": "vcenter",
        })
        if num_format:
            p["num_format"] = num_format
        return _f(wb, **p)

    def _sem_fmt(alt, num_format=None):
        p = _brd({
            "font_size": 10, "font_color": "#5C4800",
            "bg_color": C_AMARELO, "align": "center", "valign": "vcenter",
        })
        if num_format:
            p["num_format"] = num_format
        return _f(wb, **p)

    def _dif_fmt(num_format=None):
        p = _brd({
            "font_size": 10, "bg_color": C_CINZA_CLARO,
            "align": "center", "valign": "vcenter",
        })
        if num_format:
            p["num_format"] = num_format
        return _f(wb, **p)

    f_delta_pos = wb.add_format(_brd({
        "font_name": FONT_DEFAULT, "bold": True, "font_size": 10,
        "font_color": C_VERDE_ESCURO, "bg_color": C_VERDE_CLARO,
        "align": "center", "valign": "vcenter",
    }))
    f_delta_neg = wb.add_format(_brd({
        "font_name": FONT_DEFAULT, "font_size": 10,
        "font_color": C_VERMELHO, "bg_color": C_VERMELHO_CLARO,
        "align": "center", "valign": "vcenter",
    }))

    f_defs_hdr = _f(wb, **_brd_thick({
        "bold": True, "font_size": 11, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "left", "valign": "vcenter",
        "indent": 2,
    }))
    f_def_body = _f(wb, **_brd({
        "italic": True, "font_size": 9, "font_color": "#3C3C3C",
        "bg_color": C_BRANCO, "align": "left", "valign": "vcenter",
        "text_wrap": True, "indent": 2,
    }))
    f_def_body_z = _f(wb, **_brd({
        "italic": True, "font_size": 9, "font_color": "#3C3C3C",
        "bg_color": C_CINZA_CLARO, "align": "left", "valign": "vcenter",
        "text_wrap": True, "indent": 2,
    }))
    f_na = _f(wb, **_brd({
        "italic": True, "font_size": 10, "font_color": "#595959",
        "bg_color": C_CINZA_NA, "align": "center", "valign": "vcenter",
    }))

    # ── Row 0: spacer ──────────────────────────────────────────────────────
    ws.set_row(0, 8)

    # ── Row 1: Título (B:E merged, height 44) ─────────────────────────────
    ws.set_row(1, 44)
    title_text = f"DigAI & {cliente}  |  Por Grupo — Com vs Sem DigAI  |  {periodo}"
    ws.merge_range(1, 1, 1, 4, title_text, f_title)

    # ── Row 2: spacer ──────────────────────────────────────────────────────
    ws.set_row(2, 8)

    # ── Row 3: Sub-header ─────────────────────────────────────────────────
    ws.set_row(3, 26)
    ws.write(3, 1, "KPI", f_hdr_kpi)
    ws.write(3, 2, "Com DigAI", f_hdr_com)
    ws.write(3, 3, "Sem DigAI", f_hdr_sem)
    ws.write(3, 4, "Diferenca", f_hdr_dif)

    # ── KPI rows (0-indexed row 4+) ────────────────────────────────────────
    # Row 4: Total candidatos
    # Row 5: Contratados
    # Row 6: Taxa de contratacao
    # Row 7: SLA medio (se disponivel)
    # Row 8: Score medio DigAI (se disponivel)

    kpis = [
        {
            "label": "Total candidatos",
            "formula_com": f'=COUNTIF({BD("_digai_realizado", col_map)},"Sim")',
            "formula_sem": f'=COUNTIF({BD("_digai_realizado", col_map)},"N\u00e3o")',
            "diff_mode": "normal",
            "num_format": "#,##0",
        },
        {
            "label": "Contratados",
            "formula_com": (
                f'=COUNTIFS({BD("_digai_realizado", col_map)},"Sim",'
                f'{BD("status", col_map)},"Contratado*")'
            ),
            "formula_sem": (
                f'=COUNTIFS({BD("_digai_realizado", col_map)},"N\u00e3o",'
                f'{BD("status", col_map)},"Contratado*")'
            ),
            "diff_mode": "normal",
            "num_format": "#,##0",
        },
        {
            "label": "Taxa de contratacao",
            "formula_com": (
                f'=IFERROR(COUNTIFS({BD("_digai_realizado", col_map)},"Sim",'
                f'{BD("status", col_map)},"Contratado*")'
                f'/COUNTIF({BD("_digai_realizado", col_map)},"Sim"),0)'
            ),
            "formula_sem": (
                f'=IFERROR(COUNTIFS({BD("_digai_realizado", col_map)},"N\u00e3o",'
                f'{BD("status", col_map)},"Contratado*")'
                f'/COUNTIF({BD("_digai_realizado", col_map)},"N\u00e3o"),0)'
            ),
            "diff_mode": "normal",
            "num_format": "0.0%",
        },
    ]

    if has_sla:
        kpis.append({
            "label": "SLA medio (dias)",
            "formula_com": (
                f'=IFERROR(AVERAGEIFS({BD("_sla_dias", col_map)},'
                f'{BD("_digai_realizado", col_map)},"Sim",'
                f'{BD("status", col_map)},"Contratado*"),"-")'
            ),
            "formula_sem": (
                f'=IFERROR(AVERAGEIFS({BD("_sla_dias", col_map)},'
                f'{BD("_digai_realizado", col_map)},"N\u00e3o",'
                f'{BD("status", col_map)},"Contratado*"),"-")'
            ),
            "diff_mode": "reversed",
            "num_format": "0.0",
        })

    if has_score:
        kpis.append({
            "label": "Score medio DigAI",
            "formula_com": (
                f'=IFERROR(AVERAGEIF({BD("_digai_realizado", col_map)},"Sim",'
                f'{BD("score_ia", col_map)}),"-")'
            ),
            "formula_sem": '"-"',
            "diff_mode": None,
            "num_format": "0.00",
        })

    first_kpi_row = 4
    for i, kpi in enumerate(kpis):
        r = first_kpi_row + i
        ws.set_row(r, 22)
        alt = i % 2 == 1

        ws.write(r, 1, kpi["label"], _lbl_fmt(alt))
        ws.write_formula(r, 2, kpi["formula_com"], _com_fmt(alt, kpi["num_format"]))

        if kpi["formula_sem"].startswith('=') or kpi["formula_sem"].startswith('"'):
            if kpi["formula_sem"].startswith('='):
                ws.write_formula(r, 3, kpi["formula_sem"], _sem_fmt(alt, kpi["num_format"]))
            else:
                ws.write(r, 3, kpi["formula_sem"].strip('"'), f_na)
        else:
            ws.write_formula(r, 3, kpi["formula_sem"], _sem_fmt(alt, kpi["num_format"]))

        if kpi["diff_mode"] == "normal":
            ws.write_formula(r, 4, f"=C{r+1}-D{r+1}", _dif_fmt(kpi["num_format"]))
        elif kpi["diff_mode"] == "reversed":
            ws.write_formula(r, 4, f"=D{r+1}-C{r+1}", _dif_fmt(kpi["num_format"]))
        else:
            ws.write(r, 4, "\u2014", _dif_fmt())

    last_kpi_row = first_kpi_row + len(kpis) - 1

    # Conditional formatting on col E (Diferenca, index 4)
    ws.conditional_format(first_kpi_row, 4, last_kpi_row, 4, {
        "type": "cell", "criteria": ">", "value": 0, "format": f_delta_pos,
    })
    ws.conditional_format(first_kpi_row, 4, last_kpi_row, 4, {
        "type": "cell", "criteria": "<", "value": 0, "format": f_delta_neg,
    })

    # ── Definicoes ─────────────────────────────────────────────────────────
    spacer_row = last_kpi_row + 1
    ws.set_row(spacer_row, 8)

    defs_hdr_row = spacer_row + 1
    ws.set_row(defs_hdr_row, 24)
    ws.merge_range(defs_hdr_row, 1, defs_hdr_row, 4, "Definicoes", f_defs_hdr)

    definitions = [
        "Com DigAI: candidatos cuja coluna 'Passou pela DigAI' = 'Sim' (entrevista IA realizada)",
        "Sem DigAI: candidatos cuja coluna 'Passou pela DigAI' = 'Nao' (processo tradicional)",
        "Taxa de contratacao: Contratados / Total de candidatos no grupo",
        "SLA medio: media de dias da inscricao ate contratacao (menor = mais rapido)",
    ]
    for i, defi in enumerate(definitions):
        dr = defs_hdr_row + 1 + i
        ws.set_row(dr, 20)
        alt = i % 2 == 1
        df_fmt = f_def_body_z if alt else f_def_body
        ws.merge_range(dr, 1, dr, 4, defi, df_fmt)


# ─── Aba 4 (nova): Pivot Calculos ────────────────────────────────────────────

def _build_pivot_calculos(
    wb: xlsxwriter.Workbook, ws,
    col_map: dict, relatorio: dict, df: pd.DataFrame,
) -> None:
    """
    Aba 4 — Pivot Calculos: contagens e distribuicoes calculadas em Python
    (valores escritos como numeros — sem formulas cross-sheet para meses/labels dinamicos).
    """
    ws.hide_gridlines(2)
    ws.set_column(0, 0, 3)
    ws.set_column(1, 1, 32)
    ws.set_column(2, 2, 14)
    ws.set_column(3, 3, 13)
    ws.set_column(4, 4, 13)
    ws.set_column(5, 5, 20)
    ws.set_column(6, 6, 3)

    meta    = relatorio["meta"]
    cliente = meta.get("cliente", "")
    periodo = meta.get("periodo", "")

    # ── Formatos ─────────────────────────────────────────────────────────────
    f_title = _f(wb, **_brd_thick({
        "bold": True, "font_size": 14, "font_color": C_BRANCO,
        "bg_color": C_AZUL_ESCURO, "align": "left", "valign": "vcenter",
    }))
    f_sec = _f(wb, **_brd({
        "bold": True, "font_size": 11, "font_color": C_BRANCO,
        "bg_color": C_AZUL_MEDIO, "align": "left", "valign": "vcenter",
        "indent": 1,
    }))
    f_hdr = _f(wb, **_brd({
        "bold": True, "font_size": 9, "font_color": C_BRANCO,
        "bg_color": C_AZUL_MEDIO, "align": "center", "valign": "vcenter",
    }))
    f_mes_w = _f(wb, **_brd({
        "bold": True, "font_size": 10, "bg_color": C_BRANCO,
        "align": "center", "valign": "vcenter",
    }))
    f_mes_z = _f(wb, **_brd({
        "bold": True, "font_size": 10, "bg_color": C_CINZA_CLARO,
        "align": "center", "valign": "vcenter",
    }))
    f_com_v = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": C_AZUL_ESCURO,
        "bg_color": C_AZUL_CLARO, "align": "center", "valign": "vcenter",
        "num_format": "#,##0",
    }))
    f_sem_v = _f(wb, **_brd({
        "font_size": 10, "font_color": "#5C4800",
        "bg_color": C_AMARELO, "align": "center", "valign": "vcenter",
        "num_format": "#,##0",
    }))
    f_tot_w = _f(wb, **_brd({
        "font_size": 10, "bg_color": C_BRANCO, "align": "center", "valign": "vcenter",
        "num_format": "#,##0",
    }))
    f_tot_z = _f(wb, **_brd({
        "font_size": 10, "bg_color": C_CINZA_CLARO, "align": "center", "valign": "vcenter",
        "num_format": "#,##0",
    }))
    f_pct_verde = _f(wb, **_brd({
        "bold": True, "font_size": 10, "font_color": C_VERDE_ESCURO,
        "bg_color": C_VERDE_CLARO, "align": "center", "valign": "vcenter",
        "num_format": "0.0%",
    }))
    f_pct_cinza = _f(wb, **_brd({
        "font_size": 10, "font_color": "#595959",
        "bg_color": C_CINZA_CLARO, "align": "center", "valign": "vcenter",
        "num_format": "0.0%",
    }))

    # ── Row 0: spacer ──────────────────────────────────────────────────────
    ws.set_row(0, 8)

    # ── Row 1: Titulo ─────────────────────────────────────────────────────
    ws.set_row(1, 44)
    ws.merge_range(1, 1, 1, 5,
                   f"DigAI & {cliente}  |  Pivot Calculos  |  {periodo}",
                   f_title)

    row = 2

    # ═══════════════════════════════════════════════════════════════════════
    # BLOCO A — Evolucao Mensal de Contratacoes
    # ═══════════════════════════════════════════════════════════════════════
    ws.set_row(row, 8); row += 1  # spacer

    ws.set_row(row, 22)
    ws.merge_range(row, 1, row, 5, "Evolucao Mensal", f_sec)
    row += 1

    ws.set_row(row, 22)
    for j, h in enumerate(["Mes", "Com DigAI", "Sem DigAI", "Total", "% Com DigAI"], start=1):
        ws.write(row, j, h, f_hdr)
    row += 1

    evolucao_data = _compute_evolucao_mensal(df)
    for i, rd in enumerate(evolucao_data):
        ws.set_row(row, 20)
        alt = i % 2 == 1
        pct = rd["pct_digai"]

        ws.write(row, 1, rd["mes"], f_mes_z if alt else f_mes_w)
        ws.write(row, 2, rd["com_digai"], f_com_v)
        ws.write(row, 3, rd["sem_digai"], f_sem_v)
        ws.write(row, 4, rd["total"], f_tot_z if alt else f_tot_w)
        ws.write(row, 5, pct, f_pct_verde if pct >= 0.15 else f_pct_cinza)
        row += 1

    if not evolucao_data:
        ws.set_row(row, 20)
        f_na_cell = _f(wb, **_brd({
            "italic": True, "font_size": 10, "font_color": "#595959",
            "bg_color": C_CINZA_NA, "align": "center", "valign": "vcenter",
        }))
        ws.write(row, 1, "Sem dados de evolucao mensal.", f_na_cell)
        row += 1

    # ═══════════════════════════════════════════════════════════════════════
    # BLOCO B — Distribuicao por Status
    # ═══════════════════════════════════════════════════════════════════════
    ws.set_row(row, 8); row += 1  # spacer

    ws.set_row(row, 22)
    ws.merge_range(row, 1, row, 5, "Distribuicao por Status", f_sec)
    row += 1

    ws.set_row(row, 22)
    for j, h in enumerate(["Status", "Com DigAI", "% Com", "Sem DigAI", "% Sem"], start=1):
        ws.write(row, j, h, f_hdr)
    row += 1

    status_color_map = {
        "Contratado": (C_VERDE_CLARO, C_VERDE_ESCURO),
        "Reprovado":  (C_VERMELHO_CLARO, C_VERMELHO),
        "Desistiu":   (C_LARANJA_CLARO, C_LARANJA),
        "Em processo": (C_AZUL_CLARO, C_AZUL_ESCURO),
    }
    # Calcular status_counts
    if "status" in df.columns and "processo_seletivo" in df.columns:
        status_counts = (
            df.groupby(["status", "processo_seletivo"])
              .size()
              .unstack(fill_value=0)
        )
        com_col = "Com DigAI"
        sem_col = "Sem DigAI"
        com_total = status_counts.get(com_col, pd.Series(dtype=int)).sum()
        sem_total = status_counts.get(sem_col, pd.Series(dtype=int)).sum()

        for i, (st_label, st_pattern) in enumerate([
            ("Contratado", "Contratad"),
            ("Reprovado",  "Reprovad"),
            ("Desistiu",   "Desist"),
            ("Em processo", None),
        ]):
            # Aggregate matching rows
            mask = status_counts.index.to_series().astype(str)
            if st_pattern:
                matched = status_counts[mask.str.contains(st_pattern, na=False)]
            else:
                matched = status_counts[
                    ~mask.str.contains("Contratad|Reprovad|Desist", na=False)
                ]

            n_com = int(matched.get(com_col, pd.Series(dtype=int)).sum()) if com_col in matched.columns else 0
            n_sem = int(matched.get(sem_col, pd.Series(dtype=int)).sum()) if sem_col in matched.columns else 0
            pct_com = n_com / com_total if com_total > 0 else 0.0
            pct_sem = n_sem / sem_total if sem_total > 0 else 0.0

            bg, fc = status_color_map.get(st_label, (C_CINZA_CLARO, "#000000"))
            f_lbl_s = _f(wb, **_brd({
                "bold": True, "font_size": 10, "font_color": fc,
                "bg_color": bg, "align": "left", "valign": "vcenter",
            }))
            f_num_s = _f(wb, **_brd({
                "font_size": 10, "font_color": fc,
                "bg_color": bg, "align": "center", "valign": "vcenter",
                "num_format": "#,##0",
            }))
            f_pct_s = _f(wb, **_brd({
                "font_size": 10, "font_color": fc,
                "bg_color": bg, "align": "center", "valign": "vcenter",
                "num_format": "0.0%",
            }))

            ws.set_row(row, 20)
            ws.write(row, 1, st_label, f_lbl_s)
            ws.write(row, 2, n_com,    f_num_s)
            ws.write(row, 3, pct_com,  f_pct_s)
            ws.write(row, 4, n_sem,    f_num_s)
            ws.write(row, 5, pct_sem,  f_pct_s)
            row += 1
    else:
        f_na_cell = _f(wb, **_brd({
            "italic": True, "font_size": 10, "font_color": "#595959",
            "bg_color": C_CINZA_NA, "align": "center", "valign": "vcenter",
        }))
        ws.set_row(row, 20)
        ws.write(row, 1, "Sem dados de status disponíveis.", f_na_cell)
        row += 1

    # ═══════════════════════════════════════════════════════════════════════
    # BLOCO C — Top Vagas (Com DigAI)
    # ═══════════════════════════════════════════════════════════════════════
    ws.set_row(row, 8); row += 1  # spacer

    ws.set_row(row, 22)
    ws.merge_range(row, 1, row, 5, "Top Vagas — Com DigAI", f_sec)
    row += 1

    ws.set_row(row, 22)
    for j, h in enumerate(["Vaga", "Total candidatos", "Contratados", "Taxa"], start=1):
        ws.write(row, j, h, f_hdr)
    row += 1

    if "vaga" in df.columns and "processo_seletivo" in df.columns:
        com_df = df[df["processo_seletivo"].astype(str) == "Com DigAI"]
        if not com_df.empty and "status" in com_df.columns:
            top_vagas = (
                com_df.groupby("vaga").agg(
                    total=("email" if "email" in com_df.columns else "vaga", "count"),
                    contratados=("status", lambda s: int(
                        s.astype(str).str.contains("Contratad", na=False).sum()
                    )),
                )
                .sort_values("total", ascending=False)
                .head(10)
                .reset_index()
            )

            f_vaga_w = _f(wb, **_brd({
                "bold": True, "font_size": 10, "bg_color": C_BRANCO,
                "align": "left", "valign": "vcenter", "text_wrap": True,
            }))
            f_vaga_z = _f(wb, **_brd({
                "bold": True, "font_size": 10, "bg_color": C_CINZA_CLARO,
                "align": "left", "valign": "vcenter", "text_wrap": True,
            }))

            for i, vrow in top_vagas.iterrows():
                alt = i % 2 == 1
                total_v = int(vrow["total"])
                cont_v  = int(vrow["contratados"])
                taxa_v  = cont_v / total_v if total_v > 0 else 0.0

                ws.set_row(row, 20)
                ws.write(row, 1, str(vrow["vaga"]), f_vaga_z if alt else f_vaga_w)
                ws.write(row, 2, total_v, f_tot_z if alt else f_tot_w)
                ws.write(row, 3, cont_v,  f_com_v)
                ws.write(row, 4, taxa_v,  f_pct_verde if taxa_v >= 0.15 else f_pct_cinza)
                row += 1
        else:
            f_na_cell = _f(wb, **_brd({
                "italic": True, "font_size": 10, "font_color": "#595959",
                "bg_color": C_CINZA_NA, "align": "center", "valign": "vcenter",
            }))
            ws.set_row(row, 20)
            ws.write(row, 1, "Sem candidatos Com DigAI disponíveis.", f_na_cell)
            row += 1
    else:
        f_na_cell = _f(wb, **_brd({
            "italic": True, "font_size": 10, "font_color": "#595959",
            "bg_color": C_CINZA_NA, "align": "center", "valign": "vcenter",
        }))
        ws.set_row(row, 20)
        ws.write(row, 1, "Coluna 'vaga' nao disponível.", f_na_cell)
        row += 1


# ─── Aba 6: Base de Dados ─────────────────────────────────────────────────────

def _build_base_dados(
    wb: xlsxwriter.Workbook, ws,
    df: pd.DataFrame, col_map: dict,
):
    """
    Aba 6 — raw auditável.
    constant_memory=True: escrita linha a linha.
    Colorização condicional por valor (Status, Processo Seletivo, Score IA).
    """
    ws.hide_gridlines(2)
    ws.set_zoom(90)

    cols = col_map["_cols_ordered"]
    n_cols = len(cols)

    # Larguras de coluna
    for j, (src, lbl) in enumerate(cols):
        ws.set_column(j, j, max(len(lbl) + 2, 14))

    # ── Formatos de cabeçalho ──────────────────────────────────────────────
    f_hdr = wb.add_format(_brd({
        "font_name": FONT_DEFAULT, "bold": True, "font_size": 9,
        "font_color": C_BRANCO, "bg_color": C_AZUL_ESCURO,
        "align": "center", "valign": "vcenter", "text_wrap": True,
    }))

    # ── Formatos de célula base (zebra) ───────────────────────────────────
    f_cell_odd  = wb.add_format(_brd({
        "font_name": FONT_DEFAULT, "font_size": 9,
        "bg_color": C_BRANCO, "align": "left", "valign": "vcenter",
    }))
    f_cell_even = wb.add_format(_brd({
        "font_name": FONT_DEFAULT, "font_size": 9,
        "bg_color": "#F5F5F5", "align": "left", "valign": "vcenter",
    }))
    f_date_odd  = wb.add_format(_brd({
        "font_name": FONT_DEFAULT, "font_size": 9,
        "bg_color": C_BRANCO, "num_format": "dd/mm/yyyy",
    }))
    f_date_even = wb.add_format(_brd({
        "font_name": FONT_DEFAULT, "font_size": 9,
        "bg_color": "#F5F5F5", "num_format": "dd/mm/yyyy",
    }))
    f_int_odd  = wb.add_format(_brd({
        "font_name": FONT_DEFAULT, "font_size": 9,
        "bg_color": C_BRANCO, "num_format": "#,##0",
    }))
    f_int_even = wb.add_format(_brd({
        "font_name": FONT_DEFAULT, "font_size": 9,
        "bg_color": "#F5F5F5", "num_format": "#,##0",
    }))

    # ── Formatos condicionais: Status ──────────────────────────────────────
    def _status_fmt(status_val):
        if "Contratad" in status_val:
            return wb.add_format(_brd({
                "font_name": FONT_DEFAULT, "bold": True, "font_size": 9,
                "font_color": C_VERDE_ESCURO, "bg_color": C_VERDE_CLARO,
            }))
        if "Reprovad" in status_val:
            return wb.add_format(_brd({
                "font_name": FONT_DEFAULT, "font_size": 9,
                "font_color": C_VERMELHO, "bg_color": C_VERMELHO_CLARO,
            }))
        if "Desistiu" in status_val or "Desist" in status_val:
            return wb.add_format(_brd({
                "font_name": FONT_DEFAULT, "font_size": 9,
                "font_color": C_LARANJA, "bg_color": C_LARANJA_CLARO,
            }))
        return None

    def _ps_fmt(ps_val):
        if "Com DigAI" in ps_val:
            return wb.add_format(_brd({
                "font_name": FONT_DEFAULT, "bold": True, "font_size": 9,
                "font_color": C_AZUL_ESCURO, "bg_color": C_AZUL_CLARO,
            }))
        if "Sem DigAI" in ps_val:
            return wb.add_format(_brd({
                "font_name": FONT_DEFAULT, "font_size": 9,
                "font_color": "#5C4800", "bg_color": C_AMARELO,
            }))
        return None

    def _score_fmt(score_val):
        if score_val >= 7.0:
            return wb.add_format(_brd({
                "font_name": FONT_DEFAULT, "bold": True, "font_size": 9,
                "font_color": C_VERDE_ESCURO, "bg_color": C_VERDE_CLARO,
                "num_format": "0.0",
            }))
        if score_val >= 5.0:
            return wb.add_format(_brd({
                "font_name": FONT_DEFAULT, "font_size": 9,
                "font_color": C_LARANJA, "bg_color": C_AMARELO,
                "num_format": "0.0",
            }))
        return wb.add_format(_brd({
            "font_name": FONT_DEFAULT, "font_size": 9,
            "font_color": C_VERMELHO, "bg_color": C_VERMELHO_CLARO,
            "num_format": "0.0",
        }))

    # ── Cabeçalho (row 0, height 28) ──────────────────────────────────────
    ws.set_row(0, 28)
    for j, (src, lbl) in enumerate(cols):
        ws.write(0, j, lbl, f_hdr)

    ws.autofilter(0, 0, 0, n_cols - 1)
    ws.freeze_panes(1, 0)

    date_cols = {"data_cadastro", "data_ei", "data_final", "data_contratacao"}

    # ── Dados linha a linha ───────────────────────────────────────────────
    for row_idx, (_, row_data) in enumerate(df.iterrows()):
        r = row_idx + 1
        ws.set_row(r, 15)
        alt = row_idx % 2 == 1  # True = even data row

        for j, (src, _) in enumerate(cols):
            val = row_data.get(src)

            # Normalização de tipos
            if not isinstance(val, (list, dict)):
                try:
                    is_na = pd.isna(val)
                except (TypeError, ValueError):
                    is_na = False
                if is_na:
                    val = None

            if val is None:
                ws.write_blank(r, j, None, f_cell_even if alt else f_cell_odd)
                continue

            if isinstance(val, (pd.Timestamp, datetime)):
                val_d = val.date() if pd.notna(val) else None
                if val_d is None:
                    ws.write_blank(r, j, None, f_cell_even if alt else f_cell_odd)
                else:
                    ws.write_datetime(r, j, val_d, f_date_even if alt else f_date_odd)
                continue

            if isinstance(val, date) and not isinstance(val, datetime):
                if src in date_cols:
                    ws.write_datetime(r, j, val, f_date_even if alt else f_date_odd)
                else:
                    ws.write(r, j, str(val), f_cell_even if alt else f_cell_odd)
                continue

            if isinstance(val, (np.bool_,)) or isinstance(val, bool):
                val = "Sim" if val else "Nao"

            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val) if not np.isnan(val) else None
                if val is None:
                    ws.write_blank(r, j, None, f_cell_even if alt else f_cell_odd)
                    continue

            # Colorização condicional
            if src == "status" and isinstance(val, str):
                cf = _status_fmt(val)
                if cf:
                    ws.write(r, j, val, cf)
                    continue

            if src == "processo_seletivo" and isinstance(val, str):
                cf = _ps_fmt(val)
                if cf:
                    ws.write(r, j, val, cf)
                    continue

            if src == "score_ia" and isinstance(val, (int, float)):
                ws.write(r, j, val, _score_fmt(val))
                continue

            if src == "_digai_realizado" and isinstance(val, str):
                if val == "Sim":
                    cf = wb.add_format({
                        "font_name": FONT_DEFAULT, "font_size": 9, "bold": True,
                        "font_color": C_AZUL_ESCURO, "bg_color": C_AZUL_CLARO,
                        "align": "center", "border": 1, "border_color": C_BORDER_STD,
                    })
                else:
                    cf = wb.add_format({
                        "font_name": FONT_DEFAULT, "font_size": 9,
                        "font_color": "#595959", "bg_color": C_CINZA_NA,
                        "align": "center", "border": 1, "border_color": C_BORDER_STD,
                    })
                ws.write(r, j, val, cf)
                continue

            if src == "_sla_dias":
                ws.write(r, j, val, f_int_even if alt else f_int_odd)
            else:
                ws.write(r, j, val, f_cell_even if alt else f_cell_odd)


# ─── Pré-processamento ────────────────────────────────────────────────────────

def _preprocess_df_full(df: pd.DataFrame) -> pd.DataFrame:
    """Reset de índice, computa _sla_dias e _digai_realizado."""
    df = df.reset_index(drop=True)

    # SLA em dias (data_final - data_cadastro)
    if "data_final" in df.columns and "data_cadastro" in df.columns:
        df["_sla_dias"] = (
            pd.to_datetime(df["data_final"], errors="coerce") -
            pd.to_datetime(df["data_cadastro"], errors="coerce")
        ).dt.days

    # Spec LOGICA_CRUZAMENTO PASSO 9: coluna "Passou pela DigAI" = Sim/Não
    # Prioridade: processo_seletivo == "Com DigAI" (engloba email match + data_ei)
    # Fallback: _in_digai flag
    if "_digai_realizado" not in df.columns:
        if "processo_seletivo" in df.columns:
            df["_digai_realizado"] = df["processo_seletivo"].astype(str).map(
                lambda x: "Sim" if x == "Com DigAI" else "Não"
            )
        elif "_in_digai" in df.columns:
            df["_digai_realizado"] = df["_in_digai"].fillna(False).map(
                {True: "Sim", False: "Não"}
            )

    return df


# ─── Ponto de entrada público ─────────────────────────────────────────────────

def gerar_excel(
    relatorio: dict,
    params: dict,
    output_path: str,
    segmentacao_dims: list = None,
) -> str:
    """
    Gera arquivo Excel com abas per LAYOUT_SPEC.md usando xlsxwriter constant_memory.

    Parameters
    ----------
    relatorio        : dict retornado por gerar_relatorio_from_sources()
    params           : dict de parâmetros (mensalidade, salário, etc.)
    output_path      : caminho de saída (.xlsx)
    segmentacao_dims : lista de dicts {col, label} para gerar abas de segmentação.
                       Ex: [{"col": "Filial", "label": "Filial"}]

    Returns
    -------
    str — caminho absoluto do arquivo gerado
    """
    df = relatorio.pop("_df", None)
    if df is None:
        raise ValueError(
            "DataFrame interno nao encontrado no relatorio. "
            "Use gerar_relatorio_from_sources() para gerar o relatorio."
        )

    df = _preprocess_df_full(df)
    col_map = _compute_col_map(df)

    # Dados para abas dinâmicas (antes de entrar no constant_memory)
    evolucao_data = _compute_evolucao_mensal(df)
    depto_data    = _compute_por_departamento(df)

    has_score = (
        "score_ia" in df.columns and
        df["score_ia"].notna().any() and
        col_map.get("score_ia") is not None
    )

    wb = xlsxwriter.Workbook(
        output_path,
        {
            "constant_memory":    True,
            "strings_to_numbers": True,
            "default_date_format": "dd/mm/yyyy",
        },
    )

    # ── Cria abas na nova ordem: Indicadores → ROI → Por Time → Pivot → [Seg.] → Base ──
    print("   Construindo aba Indicadores DigAI...")
    ws_ind  = wb.add_worksheet(_INDICADORES_SHEET)

    print("   Construindo aba Calculadora ROI...")
    ws_roi  = wb.add_worksheet(_ROI_SHEET)

    print("   Construindo aba Por Time...")
    ws_time = wb.add_worksheet(_POR_TIME_SHEET)

    print("   Construindo aba Pivot Calculos...")
    ws_piv  = wb.add_worksheet(_PIVOT_SHEET)

    # Abas de segmentacao (antes de Base de Dados para melhor ordem de abas)
    seg_sheets = {}
    if segmentacao_dims:
        for sdim in segmentacao_dims:
            lbl        = sdim.get("label") or sdim.get("col", "Seg")
            sheet_name = f"Seg. {lbl}"[:31]
            print(f"   Construindo aba {sheet_name}...")
            seg_sheets[lbl] = wb.add_worksheet(sheet_name)

    print("   Construindo aba Base de Dados...")
    ws_base = wb.add_worksheet(_BASE_SHEET)

    # ── Preenche abas ─────────────────────────────────────────────────────
    _build_indicadores(wb, ws_ind, col_map, relatorio, params, df)
    _build_calculadora_roi(wb, ws_roi, col_map, relatorio, params)
    _build_por_time(wb, ws_time, col_map, relatorio, df)
    _build_pivot_calculos(wb, ws_piv, col_map, relatorio, df)

    # Abas de segmentacao
    if segmentacao_dims:
        for sdim in segmentacao_dims:
            lbl     = sdim.get("label") or sdim.get("col", "Seg")
            ws_seg  = seg_sheets.get(lbl)
            dim_col = sdim.get("col", "")
            if ws_seg is None:
                continue
            if dim_col in df.columns:
                seg_data = _compute_segmentacao_data(df, dim_col)
                _build_segmentacao_tab(wb, ws_seg, seg_data, lbl)
            else:
                ws_seg.hide_gridlines(2)
                ws_seg.write(0, 0, f"Coluna '{dim_col}' nao encontrada.")

    # Base de Dados — SEMPRE por ultimo (constant_memory: row-by-row)
    _build_base_dados(wb, ws_base, df, col_map)

    # Libera DataFrame antes de fechar
    del df
    gc.collect()

    wb.close()
    print(f"   Excel salvo: {output_path}")
    return output_path


# ─── Compatibilidade com excel_segmented.py (openpyxl) ───────────────────────

def _build_base(wb_openpyxl, df: pd.DataFrame) -> dict:
    """
    Stub openpyxl para excel_segmented.py.
    Cria aba 'Base de Dados' no workbook openpyxl e retorna col_map {src: letter}.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter as _gcl
        from openpyxl.styles import Font as _Font, PatternFill as _PFill
        from openpyxl.styles import Alignment as _Align, Border as _Border, Side as _Side
        from openpyxl.worksheet.table import Table as _Table, TableStyleInfo as _TSI
    except ImportError:
        raise ImportError("openpyxl necessario para excel_segmented.")

    def _ofill(hex_color):
        return _PFill("solid", fgColor=hex_color.lstrip("#"))

    def _ofont(bold=False, size=11, color="000000", italic=False):
        return _Font(name=FONT_DEFAULT, bold=bold, size=size,
                     color=color.lstrip("#"), italic=italic)

    def _oalign(h="left", v="center", wrap=False):
        return _Align(horizontal=h, vertical=v, wrap_text=wrap)

    def _oborder():
        s = _Side(style="thin", color=C_BORDER_STD.lstrip("#"))
        return _Border(left=s, right=s, top=s, bottom=s)

    ws = wb_openpyxl.create_sheet("Base de Dados")
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    df = df.reset_index(drop=True)
    if "data_final" in df.columns and "data_cadastro" in df.columns:
        df["_sla_dias"] = (
            pd.to_datetime(df["data_final"], errors="coerce") -
            pd.to_datetime(df["data_cadastro"], errors="coerce")
        ).dt.days

    canonical = list(_CANONICAL)
    stage_cols = df.attrs.get("stage_cols", {})
    for n in sorted(stage_cols.keys()):
        for key, lsuf in [("name", "Nome"), ("entry", "Entrada"),
                          ("exit", "Saida"), ("days", "Dias")]:
            cs = f"stage_{n}_{key}"
            if cs in df.columns:
                canonical.append((cs, f"Etapa {n} {lsuf}"))

    cols = [(s, l) for s, l in canonical if s in df.columns]

    for j, (src, lbl) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=lbl)
        c.fill   = _ofill(C_AZUL_ESCURO)
        c.font   = _ofont(bold=True, size=9, color=C_BRANCO)
        c.alignment = _oalign("center")
        c.border = _oborder()
        ws.column_dimensions[_gcl(j)].width = max(len(lbl) + 2, 14)
    ws.row_dimensions[1].height = 28

    date_cols_set = {"data_cadastro", "data_ei", "data_final", "data_contratacao"}
    for row_idx, (_, row_data) in enumerate(df.iterrows()):
        r = row_idx + 2
        bg_hex = "F5F5F5" if row_idx % 2 == 1 else "FFFFFF"
        for j, (src, _) in enumerate(cols, start=1):
            val = row_data.get(src)
            try:
                is_na = pd.isna(val) if not isinstance(val, (list, dict)) else False
            except (TypeError, ValueError):
                is_na = False
            if is_na:
                val = None
            elif isinstance(val, (pd.Timestamp, datetime)):
                val = val.date() if pd.notna(val) else None
            elif isinstance(val, (np.bool_,)) or isinstance(val, bool):
                val = "Sim" if val else "Nao"
            elif isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val) if not np.isnan(val) else None

            c = ws.cell(row=r, column=j, value=val)
            c.font   = _ofont(size=9)
            c.fill   = _ofill(bg_hex)
            c.border = _oborder()
            if src in date_cols_set and val is not None:
                c.number_format = "DD/MM/YYYY"
            elif src == "_sla_dias" and val is not None:
                c.number_format = "#,##0"

    total_rows = len(df)
    total_cols = len(cols)
    last_col_letter = _gcl(total_cols)
    table_ref = f"A1:{last_col_letter}{total_rows + 1}"
    table = _Table(displayName="BaseDados", ref=table_ref)
    table.tableStyleInfo = _TSI(
        name="TableStyleMedium9",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"

    return {src: _gcl(j) for j, (src, _) in enumerate(cols, start=1)}


def _build_duplicatas(wb_openpyxl, df: pd.DataFrame) -> None:
    """Stub mantido para compatibilidade. Não é chamado em produção (REGRA 1)."""
    pass
