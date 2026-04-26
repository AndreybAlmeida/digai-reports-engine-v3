"""
gerar_relatorio.py — EXCEL_LAYOUT_DEFINITIVO
=============================================
5 abas obrigatórias (em ordem):
  1. Indicadores DigAI    — KPIs com fórmulas Excel apontando para Base de Dados
  2. Calculadora ROI      — inputs editáveis + fórmulas de saving/ROI
  3. Por Time             — segmentação por unidade de negócio (empresa)
  4. Pivot Cálculos       — score por faixa, distribuição por área, evolução mensal
  5. Base de Dados        — 12 colunas fixas, criada por último na ordem visual

2 abas opcionais:
  6. Ranking por Empresa  — quando há ≥ 3 valores distintos de empresa
  7. Origem dos Candidatos — quando phone ≥ 50% preenchido
"""

import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta ────────────────────────────────────────────────────────────────────
C = {
    'AZUL_ESC':  '1F3864',
    'AZUL_MED':  '2E75B6',
    'AZUL_CL':   'BDD7EE',
    'VERDE':     '375623',
    'VERDE_L':   'E2EFDA',
    'AMARELO':   'FFF2CC',
    'AMAR_BD':   'FFD966',
    'LARANJA':   'ED7D31',
    'LARANJA_L': 'FCE4D6',
    'CINZA':     'F2F2F2',
    'BRANCO':    'FFFFFF',
    'VERM':      'C00000',
    'VERM_L':    'FFE7E7',
    'CINZA_NA':  'D9D9D9',
    'ROXO':      '7030A0',
    'ROXO_L':    'E2D4F0',
}

# ── DDD → Estado (Brasil completo) ────────────────────────────────────────────
DDD_ESTADO = {
    '11': 'SP', '12': 'SP', '13': 'SP', '14': 'SP', '15': 'SP',
    '16': 'SP', '17': 'SP', '18': 'SP', '19': 'SP',
    '21': 'RJ', '22': 'RJ', '24': 'RJ',
    '27': 'ES', '28': 'ES',
    '31': 'MG', '32': 'MG', '33': 'MG', '34': 'MG', '35': 'MG',
    '37': 'MG', '38': 'MG',
    '41': 'PR', '42': 'PR', '43': 'PR', '44': 'PR', '45': 'PR', '46': 'PR',
    '47': 'SC', '48': 'SC', '49': 'SC',
    '51': 'RS', '53': 'RS', '54': 'RS', '55': 'RS',
    '61': 'DF', '62': 'GO', '63': 'TO', '64': 'GO',
    '65': 'MT', '66': 'MT',
    '67': 'MS',
    '68': 'AC',
    '69': 'RO',
    '71': 'BA', '73': 'BA', '74': 'BA', '75': 'BA', '77': 'BA',
    '79': 'SE',
    '81': 'PE', '82': 'AL', '83': 'PB', '84': 'RN', '85': 'CE',
    '86': 'PI', '87': 'PE', '88': 'CE', '89': 'PI',
    '91': 'PA', '92': 'AM', '93': 'PA', '94': 'PA',
    '95': 'RR', '96': 'AP', '97': 'AM', '98': 'MA', '99': 'MA',
}

ESTADO_NOME = {
    'AC': 'Acre',       'AL': 'Alagoas',       'AP': 'Amapá',
    'AM': 'Amazonas',   'BA': 'Bahia',         'CE': 'Ceará',
    'DF': 'Dist. Fed.', 'ES': 'Espírito Santo', 'GO': 'Goiás',
    'MA': 'Maranhão',   'MT': 'Mato Grosso',   'MS': 'Mato Grosso do Sul',
    'MG': 'Minas Gerais','PA': 'Pará',         'PB': 'Paraíba',
    'PR': 'Paraná',     'PE': 'Pernambuco',    'PI': 'Piauí',
    'RJ': 'Rio de Janeiro','RN': 'Rio Grande do Norte','RS': 'Rio Grande do Sul',
    'RO': 'Rondônia',   'RR': 'Roraima',       'SC': 'Santa Catarina',
    'SP': 'São Paulo',  'SE': 'Sergipe',       'TO': 'Tocantins',
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def brd(cell, cor='9DC3E6'):
    s = Side(style='thin', color=cor)
    cell.border = Border(left=s, right=s, top=s, bottom=s)


def brd_thick(cell):
    s = Side(style='medium', color='1F3864')
    cell.border = Border(left=s, right=s, top=s, bottom=s)


def sc(ws, r, c, v=None, bg='FFFFFF', fg='212121', bold=False, sz=10,
       ha='left', wrap=False, fmt=None, italic=False, ind=0):
    cl = ws.cell(r, c)
    if v is not None:
        cl.value = v
    cl.font      = Font(name='Calibri', size=sz, bold=bold, color=fg, italic=italic)
    cl.fill      = PatternFill('solid', start_color=bg)
    cl.alignment = Alignment(horizontal=ha, vertical='center',
                             wrap_text=wrap, indent=ind)
    if fmt:
        cl.number_format = fmt
    brd(cl)
    return cl


def inp(ws, r, c, v, fmt=None):
    """Célula editável — fundo amarelo com borda dourada."""
    cl = ws.cell(r, c)
    cl.value     = v
    cl.font      = Font(name='Calibri', size=10, color='000000')
    cl.fill      = PatternFill('solid', start_color=C['AMARELO'])
    cl.alignment = Alignment(horizontal='right', vertical='center')
    if fmt:
        cl.number_format = fmt
    s = Side(style='thin', color=C['AMAR_BD'])
    cl.border = Border(left=s, right=s, top=s, bottom=s)
    return cl


def rh(ws, r, h=20):
    ws.row_dimensions[r].height = h


def cw(ws, col, w):
    ws.column_dimensions[col].width = w


def mg(ws, r, c1, c2):
    ws.merge_cells(f'{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}')


def titulo_aba(ws, r, c1, c2, texto, cliente, sz=16):
    mg(ws, r, c1, c2)
    cl = ws.cell(r, c1)
    cl.value     = f'DigAI  ×  {cliente}   |   {texto}'
    cl.font      = Font(name='Calibri', bold=True, size=sz, color=C['BRANCO'])
    cl.fill      = PatternFill('solid', start_color=C['AZUL_ESC'])
    cl.alignment = Alignment(horizontal='center', vertical='center')
    brd_thick(cl)
    rh(ws, r, 44)


def header_bloco(ws, r, c1, c2, texto, bg=None):
    bg = bg or C['AZUL_ESC']
    mg(ws, r, c1, c2)
    cl = ws.cell(r, c1)
    cl.value     = texto
    cl.font      = Font(name='Calibri', bold=True, size=11, color=C['BRANCO'])
    cl.fill      = PatternFill('solid', start_color=bg)
    cl.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    brd(cl)
    rh(ws, r, 26)


def sub_header(ws, r, cols_txts: list):
    rh(ws, r, 22)
    for col, txt in cols_txts:
        sc(ws, r, col, txt, bg=C['AZUL_MED'], fg=C['BRANCO'], bold=True, sz=9, ha='center')


def zebra(i):
    return C['CINZA'] if i % 2 == 0 else C['BRANCO']


# ── _build_base_df ────────────────────────────────────────────────────────────

def _build_base_df(df_raw):
    """
    Inclui TODOS os candidatos do ATS com flag DigAI Realizado (Sim/Não).
    12 colunas de dados + coluna M (DigAI Realizado) para auditoria e filtro.
    """
    if df_raw is None or len(df_raw) == 0:
        return pd.DataFrame()

    df = df_raw.copy()

    if len(df) == 0:
        return df

    # Máscara DigAI — usa processo_seletivo primeiro (mesma lógica do dashboard),
    # pois com_mask = _in_digai | data_ei.notna() e _in_digai só captura email/phone.
    # Usando _in_digai diretamente subestima o total de EIs vs o dashboard.
    if 'processo_seletivo' in df.columns:
        digai_mask = df['processo_seletivo'] == 'Com DigAI'
    elif '_in_digai' in df.columns:
        digai_mask = df['_in_digai'].astype(bool)
    else:
        digai_mask = pd.Series(True, index=df.index)

    def _to_str(series):
        """Converte qualquer Series (incluindo Categorical) para str com fillna('')."""
        return series.astype(object).fillna('').astype(str)

    # ── Time (A) — empresa normalizada
    if 'empresa' in df.columns:
        df['_time'] = _to_str(df['empresa']).str.strip()
    else:
        df['_time'] = ''

    # ── Área (B)
    area_found = False
    for ac in ('workspace', 'area_cand', 'departamento', 'Departamento', 'area'):
        if ac in df.columns:
            df['_area'] = _to_str(df[ac]).str.strip()
            area_found = True
            break
    if not area_found:
        df['_area'] = ''

    # ── Nome (C)
    for nc in ('nome', 'Nome', 'nome_completo', 'candidate_name'):
        if nc in df.columns:
            df['_nome'] = _to_str(df[nc])
            break
    else:
        df['_nome'] = ''

    # ── Email (D)
    df['_email'] = _to_str(df.get('email', pd.Series('', index=df.index)))

    # ── Vaga (E) — prefere vaga_digai
    if 'vaga_digai' in df.columns:
        df['_vaga'] = _to_str(df['vaga_digai'])
        if 'vaga' in df.columns:
            mask = df['_vaga'].eq('')
            df.loc[mask, '_vaga'] = _to_str(df.loc[mask, 'vaga'])
    elif 'vaga' in df.columns:
        df['_vaga'] = _to_str(df['vaga'])
    else:
        df['_vaga'] = ''

    # ── Status Vaga (F)  — status é CategoricalDtype em build_unified → usa _to_str
    for sc_col in ('status', 'Status', 'status_ia'):
        if sc_col in df.columns:
            df['_status'] = _to_str(df[sc_col])
            break
    else:
        df['_status'] = ''

    # ── Score IA (G)
    if 'score_ia' in df.columns:
        df['_score_ia'] = pd.to_numeric(df['score_ia'], errors='coerce')
    else:
        df['_score_ia'] = np.nan

    # ── Aprovado IA (H) → Sim/Não  (vazio para não-DigAI)
    if 'aprovado_ia' in df.columns:
        aprov_raw = df['aprovado_ia'].apply(
            lambda v: 'Sim' if str(v).lower() in ('true', '1', 'sim', 'yes') else 'Não'
        )
        # Candidatos não-DigAI não têm dados de aprovação
        df['_aprov'] = aprov_raw.where(digai_mask, '')
    else:
        df['_aprov'] = ''

    # ── Score Editado (I)
    if 'score_editado' in df.columns:
        df['_score_ed'] = pd.to_numeric(df['score_editado'], errors='coerce')
    else:
        df['_score_ed'] = np.nan

    # ── Req. Atendido (J) → Sim/Não
    if 'req_atendido' in df.columns:
        df['_req'] = df['req_atendido'].apply(
            lambda v: ('Sim' if str(v).lower() in ('true', '1', 'sim', 'yes')
                       else ('Não' if str(v).lower() in ('false', '0', 'não', 'nao', 'no')
                             else ''))
        )
    else:
        df['_req'] = ''

    # ── Dias Vaga Aberta (K)
    if 'dias_vaga' in df.columns:
        df['_dias_vaga'] = pd.to_numeric(df['dias_vaga'], errors='coerce')
    else:
        df['_dias_vaga'] = np.nan

    # ── Data Aplicação (L) — remove timezone para evitar erro em to_period
    for dc in ('data_ei', 'data_cadastro'):
        if dc in df.columns:
            dt = pd.to_datetime(df[dc], errors='coerce')
            if hasattr(dt, 'dt') and dt.dt.tz is not None:
                dt = dt.dt.tz_convert('UTC').dt.tz_localize(None)
            df['_data_aplic'] = dt
            break
    else:
        df['_data_aplic'] = pd.NaT

    # ── DigAI Realizado (M) — flag de auditoria: candidato passou pela EI DigAI?
    df['_digai_realizado'] = np.where(digai_mask, 'Sim', 'Não')

    return df


# ── Aba 1 — INDICADORES DIGAI ─────────────────────────────────────────────────

def aba_indicadores(wb, df_d, params, cliente, periodo=''):
    ws = wb.active
    ws.title = 'Indicadores DigAI'
    ws.sheet_view.showGridLines = False

    for col, w in [('A', 3), ('B', 34), ('C', 20), ('D', 42), ('E', 18), ('F', 40), ('G', 3)]:
        cw(ws, col, w)

    rh(ws, 1, 8)
    titulo_aba(ws, 2, 2, 6, 'Indicadores de Performance DigAI', cliente)
    rh(ws, 3, 6)

    if periodo:
        mg(ws, 4, 2, 6)
        cl = ws.cell(4, 2)
        cl.value     = f'Período: {periodo}'
        cl.font      = Font(name='Calibri', size=10, italic=True, color='595959')
        cl.fill      = PatternFill('solid', start_color=C['CINZA'])
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)
        rh(ws, 4, 18)
        rh(ws, 5, 8)
        r_kpi_start = 7
    else:
        rh(ws, 4, 8)
        r_kpi_start = 6

    # Sub-header KPI
    header_bloco(ws, r_kpi_start - 1, 2, 6, 'KPIs Calculados da Base DigAI')
    sub_header(ws, r_kpi_start, [
        (2, 'Indicador'), (3, 'Valor'), (4, 'Fórmula / Fonte'), (5, 'Status'), (6, 'Observação')
    ])

    # Filtro DigAI-only para detecção de dados e cálculos Python
    # (df_d agora tem TODOS os candidatos ATS; usamos df_c para KPIs DigAI)
    if '_digai_realizado' in df_d.columns:
        df_c = df_d[df_d['_digai_realizado'] == 'Sim']
    else:
        df_c = df_d

    # Detectar presença de dados — usa apenas candidatos DigAI
    has_aprov   = '_aprov'    in df_c.columns and df_c['_aprov'].eq('Sim').any()
    has_score   = '_score_ia' in df_c.columns and df_c['_score_ia'].notna().any()
    has_score_e = '_score_ed' in df_c.columns and df_c['_score_ed'].notna().any()
    has_status  = '_status'   in df_c.columns and df_c['_status'].ne('').any()
    has_req     = '_req'      in df_c.columns and df_c['_req'].ne('').any()
    has_dias    = '_dias_vaga' in df_c.columns and df_c['_dias_vaga'].notna().any()
    has_time    = '_time'     in df_c.columns and df_c['_time'].ne('').any()

    CALC  = '✓ Calculado'
    ESTIM = '~ Estimativa'
    ATS   = 'Aguarda ATS'

    # Calcular mensalidade e total_ei para Saving/ROI (usa apenas EIs DigAI)
    mensalidade = params.get('mensalidade', 0) or params.get('mensalidade_digai', 7600)
    max_ta      = params.get('prod_max', 0) or params.get('max_entrevistas_ta', 127)
    salario_clt = params.get('salario_clt', 0) or params.get('salario_ta_clt', 4750)

    # Preferência: usar valores pré-computados do analytics para garantir
    # consistência exata com o dashboard (mesma fonte de verdade)
    if params.get('saving') is not None:
        saving   = params['saving']
        roi      = params.get('roi') or 0
        total_ei = (params.get('total_ei') or
                    (int(df_c['_email'].ne('').sum()) if '_email' in df_c.columns else len(df_c)))
    else:
        n_meses  = params.get('n_meses', 1)
        total_ei = int(df_c['_email'].ne('').sum()) if '_email' in df_c.columns else len(df_c)
        custo_ta = salario_clt / max_ta if max_ta > 0 else 0
        custo_ia = mensalidade / total_ei if total_ei > 0 else 0
        saving   = max(0, (custo_ta - custo_ia) * total_ei)
        investimento = mensalidade * n_meses
        roi      = saving / investimento if investimento > 0 else 0

    # Detecção de times para últimas linhas (apenas DigAI)
    time_vals = []
    if has_time:
        tvs = [t for t in df_c['_time'].dropna().unique()
               if str(t).strip() not in ('', 'nan')]
        time_vals = sorted(set(str(t).strip() for t in tvs))[:8]

    # Coluna M = 'DigAI Realizado'; fórmulas usam M:M,"Sim" como filtro DigAI
    M = "'Base de Dados'!M:M,\"Sim\""

    # KPI rows: (label, formula, fmt, status, observação)
    kpi_rows = [
        (
            'Total Entrevistas DigAI',
            f"=COUNTIF({M})",
            '#,##0', CALC,
            'Candidatos que passaram pela Entrevista Inteligente DigAI'
        ),
        (
            'Aprovados pela IA',
            "=COUNTIF('Base de Dados'!H:H,\"Sim\")" if has_aprov else '—',
            '#,##0', CALC if has_aprov else ATS,
            'Candidatos com Aprovado IA = Sim'
        ),
        (
            'Taxa de Aprovação IA (Adesão)',
            f"=IFERROR(COUNTIF('Base de Dados'!H:H,\"Sim\")/COUNTIF({M}),0)"
            if has_aprov else '—',
            '0.00%', CALC if has_aprov else ATS,
            '% aprovados / total entrevistados pela IA'
        ),
        (
            'Score Médio IA',
            "=IFERROR(AVERAGE('Base de Dados'!G:G),\"—\")" if has_score else '—',
            '0.0', CALC if has_score else ATS,
            'Média do Score IA (apenas candidatos com EI DigAI)'
        ),
        (
            'Score Médio Editado',
            "=IFERROR(AVERAGE('Base de Dados'!I:I),\"—\")" if has_score_e else '—',
            '0.0', CALC if has_score_e else ATS,
            'Média do Score após edição manual pelo recrutador'
        ),
        (
            'Contratados via DigAI',
            f"=COUNTIFS('Base de Dados'!F:F,\"Contratado\",{M})" if has_status else '—',
            '#,##0', CALC if has_status else ATS,
            'Candidatos com Status Vaga = Contratado e DigAI = Sim'
        ),
        (
            'Assertividade (Taxa de Contratação)',
            f"=IFERROR(COUNTIFS('Base de Dados'!F:F,\"Contratado\",{M})/COUNTIF({M}),0)"
            if has_status else '—',
            '0.00%', CALC if has_status else ATS,
            'Contratados DigAI / Total entrevistados pela IA'
        ),
        (
            'Requisitos Atendidos',
            "=COUNTIF('Base de Dados'!J:J,\"Sim\")" if has_req else '—',
            '#,##0', CALC if has_req else ATS,
            'Candidatos com Req. Atendido = Sim'
        ),
        (
            '% Requisitos Atendidos',
            f"=IFERROR(COUNTIF('Base de Dados'!J:J,\"Sim\")/COUNTIF({M}),0)"
            if has_req else '—',
            '0.00%', CALC if has_req else ATS,
            '% candidatos que atendem os requisitos da vaga'
        ),
        (
            'Dias Médios — Vaga Aberta',
            "=IFERROR(AVERAGE('Base de Dados'!K:K),\"—\")" if has_dias else '—',
            '0.0', CALC if has_dias else ATS,
            'Média de dias com a vaga aberta (daysStayOpen)'
        ),
        (
            'Saving DigAI',
            saving,
            'R$ #,##0.00', ESTIM,
            f'(Custo EI humana − Custo EI IA) × {total_ei:,} EIs realizadas'
        ),
        (
            'ROI DigAI',
            roi,
            '0.0"x"', ESTIM,
            f'Saving / Mensalidade DigAI  (mensalidade R$ {mensalidade:,.0f})'
        ),
    ]

    # Adicionar linhas por Time (até 4)
    for tv in time_vals[:4]:
        kpi_rows.append((
            f'Entrevistas — {tv}',
            f"=COUNTIFS('Base de Dados'!A:A,\"{tv}\",{M})",
            '#,##0', CALC if has_time else ATS,
            f'Total de EIs DigAI do time / BU "{tv}"'
        ))

    # Preencher KPI table
    r = r_kpi_start + 1
    for i, (label, formula, fmt, status, obs) in enumerate(kpi_rows):
        rh(ws, r, 22)
        bg = zebra(i)

        # Status color
        if status == CALC:
            st_bg, st_fg = C['VERDE_L'], C['VERDE']
        elif status == ESTIM:
            st_bg, st_fg = C['AMARELO'], '5C4800'
        else:
            st_bg, st_fg = C['CINZA_NA'], '595959'

        sc(ws, r, 2, label, bg=bg, bold=True, sz=10)

        cl_val = sc(ws, r, 3, bg=C['AZUL_CL'], fg=C['AZUL_ESC'], bold=True, ha='right', fmt=fmt)
        if isinstance(formula, str) and formula.startswith('='):
            cl_val.value = formula
        elif formula == '—':
            cl_val.value = '—'
            cl_val.fill  = PatternFill('solid', start_color=C['CINZA_NA'])
        else:
            cl_val.value = formula  # numeric Python value

        # Fórmula / Fonte (col D) — strip '=' para mostrar texto, não executar fórmula
        if isinstance(formula, str) and formula.startswith('='):
            formula_txt = formula[1:]          # remove '=' → exibe como texto em D
        elif isinstance(formula, str):
            formula_txt = formula
        else:
            formula_txt = f'Python: {formula:.4g}'
        sc(ws, r, 4, formula_txt, bg=bg, fg='595959', italic=True, sz=8, wrap=True)

        sc(ws, r, 5, status, bg=st_bg, fg=st_fg, bold=True, sz=9, ha='center')
        sc(ws, r, 6, obs, bg=bg, fg='595959', italic=True, sz=9, wrap=True)

        r += 1

    rh(ws, r, 10)
    r += 1

    # Bloco definições
    header_bloco(ws, r, 2, 6, 'Legenda de Status', bg=C['AZUL_MED'])
    r += 1

    defs = [
        ('✓ Calculado',   C['VERDE_L'],  C['VERDE'],   'Calculado diretamente da Base de Dados via fórmula Excel'),
        ('~ Estimativa',  C['AMARELO'],  '5C4800',     'Estimativa com base nos parâmetros do cliente (editáveis na aba Calculadora ROI)'),
        ('Aguarda ATS',   C['CINZA_NA'], '595959',     'Dado não encontrado na base enviada — verifique se a coluna existe no arquivo DigAI'),
    ]
    for st_label, bg, fg, desc in defs:
        rh(ws, r, 20)
        sc(ws, r, 2, st_label, bg=bg, fg=fg, bold=True, sz=9, ha='center')
        mg(ws, r, 3, 6)
        cl = ws.cell(r, 3)
        cl.value     = desc
        cl.font      = Font(name='Calibri', size=9, color='3C3C3C')
        cl.fill      = PatternFill('solid', start_color=C['BRANCO'])
        cl.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        brd(cl)
        r += 1


# ── Aba 2 — CALCULADORA ROI ───────────────────────────────────────────────────

def aba_calculadora(wb, params, cliente):
    ws = wb.create_sheet('Calculadora ROI')
    ws.sheet_view.showGridLines = False

    for col, w in [('A', 3), ('B', 36), ('C', 22), ('D', 36), ('E', 3)]:
        cw(ws, col, w)

    mensalidade = params.get('mensalidade', 0) or params.get('mensalidade_digai', 7600)
    salario_clt = params.get('salario_clt', 0) or params.get('salario_ta_clt', 4750)
    salario_pj  = params.get('salario_pj', 0)  or params.get('salario_ta_pj', 2500)
    max_ta      = params.get('prod_max', 0)    or params.get('max_entrevistas_ta', 127)
    prod_pct    = params.get('prod_pct', 0)    or params.get('produtividade_pct', 0.60)
    tempo_ei    = params.get('tempo_entrev', 0) or params.get('tempo_entrevista_min', 30)

    rh(ws, 1, 8)
    titulo_aba(ws, 2, 2, 4, 'Calculadora de ROI', cliente, sz=14)
    rh(ws, 3, 8)

    # ── Bloco Inputs ──────────────────────────────────────────────────────────
    header_bloco(ws, 4, 2, 4, '1. Parâmetros do Cliente  [ células amarelas são editáveis ]')
    rh(ws, 5, 22)
    for col, txt in [(2, 'Dado Bruto — Mensal'), (3, 'Valor'), (4, 'Observações')]:
        sc(ws, 5, col, txt, bg=C['AZUL_MED'], fg=C['BRANCO'], bold=True,
           ha='center' if col > 2 else 'left')

    input_rows = [
        # (row_offset, label, value, fmt, obs, is_formula)
        (0, 'Salário base TA | CLT',        salario_clt, 'R$ #,##0.00',
         'Salário bruto mensal do Analista de Seleção (CLT)', False),
        (1, 'Salário base TA | PJ',         salario_pj,  'R$ #,##0.00',
         'Referência para contratos PJ', False),
        (2, 'Tempo por Entrevista (min)',    tempo_ei,    '0',
         'Duração média de uma Entrevista Inteligente', False),
        (3, 'Entrevistas feitas pela IA',
         "=COUNTIF('Base de Dados'!M:M,\"Sim\")",   '#,##0',
         'Candidatos com DigAI Realizado = Sim na Base de Dados', True),
        (4, 'Entrevistas máx TA/mês',        max_ta,      '#,##0',
         'Capacidade real do recrutador (entrevistas contratadas/mês)', False),
        (5, 'Mensalidade DigAI',             mensalidade, 'R$ #,##0.00',
         'Valor contratado mensalmente com a DigAI', False),
        (6, 'Produtividade TA',              prod_pct,    '0%',
         '% do tempo produtivo dedicado a Entrevistas', False),
    ]

    for off, label, val, fmt, obs, is_formula in input_rows:
        row = 6 + off
        rh(ws, row, 22)
        bg_l = zebra(off)
        sc(ws, row, 2, label, bg=bg_l, bold=True)
        if is_formula:
            cl = sc(ws, row, 3, bg=C['AZUL_CL'], fg=C['AZUL_ESC'], bold=True, ha='right', fmt=fmt)
            cl.value = val
        else:
            inp(ws, row, 3, val, fmt=fmt)
        sc(ws, row, 4, obs, bg=bg_l, fg='595959', italic=True, sz=9)

    rh(ws, 13, 10)

    # ── Bloco Savings ─────────────────────────────────────────────────────────
    # C6=salario_clt, C7=salario_pj, C8=tempo, C9=EI_formula, C10=max_ta, C11=mensalidade, C12=prod_pct
    header_bloco(ws, 14, 2, 4, '2. Savings DigAI  [ calculados automaticamente ]',
                 bg=C['VERDE'])
    rh(ws, 15, 32)
    sc(ws, 15, 2, 'Saving DigAI', bg=C['VERDE'], fg=C['BRANCO'], bold=True, sz=11)
    cl_sv = sc(ws, 15, 3, bg=C['VERDE'], fg=C['BRANCO'], bold=True,
               sz=14, ha='center', fmt='R$ #,##0.00')
    cl_sv.value = '=IFERROR((C6/C10-C11/C9)*C9,0)'
    sc(ws, 15, 4, '(Custo EI humana − Custo EI IA) × Qtd EIs',
       bg=C['VERDE'], fg=C['BRANCO'], sz=9, italic=True)

    n_meses_calc = int(params.get('n_meses', 1)) or 1
    rh(ws, 16, 32)
    sc(ws, 16, 2, 'ROI DigAI', bg=C['AZUL_ESC'], fg=C['BRANCO'], bold=True, sz=11)
    cl_roi = sc(ws, 16, 3, bg=C['AZUL_ESC'], fg=C['BRANCO'], bold=True,
                sz=14, ha='center', fmt='0.0"x"')
    cl_roi.value = f'=IFERROR(C15/(C11*{n_meses_calc}),0)'
    sc(ws, 16, 4, f'Saving / (Mensalidade × {n_meses_calc} meses do período)',
       bg=C['AZUL_ESC'], fg=C['BRANCO'], sz=9, italic=True)

    rh(ws, 17, 10)

    # ── Bloco Calculado ───────────────────────────────────────────────────────
    header_bloco(ws, 18, 2, 4, '3. Valores Derivados  [ referência ]',
                 bg=C['AZUL_MED'])
    rh(ws, 19, 22)
    for col, txt in [(2, 'Métrica'), (3, 'Valor'), (4, 'Observações')]:
        sc(ws, 19, col, txt, bg=C['AZUL_MED'], fg=C['BRANCO'], bold=True,
           ha='center' if col > 2 else 'left')

    derived = [
        ('Horas trabalhadas/mês',     176,              '0',           'Base CLT (22 dias × 8h)'),
        ('Horas em entrevistas',       '=C20*C12',       '0.0',         'Horas produtivas dedicadas a EIs'),
        ('Minutos em entrevistas',     '=C21*60',        '#,##0',       'Total de minutos em EIs/mês'),
        ('Qtd entrevistas/mês (TA)',   '=IFERROR(C22/C8,0)', '#,##0',  'Capacidade real do recrutador'),
        ('Custo por EI (TA)',          '=IFERROR(C6/C10,0)','R$ #,##0.00','Custo unitário por EI humana'),
        ('Custo por EI (IA)',          '=IFERROR(C11/C9,0)','R$ #,##0.00','Custo unitário por EI DigAI'),
        ('Economia por EI',            '=IFERROR(C24-C25,0)','R$ #,##0.00','Saving por entrevista substituída'),
    ]

    for i, (label, val, fmt, obs) in enumerate(derived):
        row = 20 + i
        rh(ws, row, 22)
        bg_d = zebra(i)
        sc(ws, row, 2, label, bg=bg_d, bold=True)
        cl = sc(ws, row, 3, bg=C['AZUL_CL'], fg=C['AZUL_ESC'], bold=True, ha='right', fmt=fmt)
        cl.value = val
        sc(ws, row, 4, obs, bg=bg_d, fg='595959', italic=True, sz=9)

    # Nota
    nota_r = 20 + len(derived) + 1
    rh(ws, nota_r, 36)
    mg(ws, nota_r, 2, 4)
    cl = ws.cell(nota_r, 2)
    cl.value     = ('Células em amarelo são editáveis. As demais são calculadas automaticamente '
                    'a partir da Base de Dados e dos parâmetros acima.')
    cl.font      = Font(name='Calibri', size=9, italic=True, color='4A4A4A')
    cl.fill      = PatternFill('solid', start_color=C['AMARELO'])
    cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)
    brd(cl)


# ── Aba 3 — POR TIME ──────────────────────────────────────────────────────────

def aba_por_time(wb, df_d, cliente):
    ws = wb.create_sheet('Por Time')
    ws.sheet_view.showGridLines = False

    for col, w in [('A', 3), ('B', 26), ('C', 14), ('D', 14), ('E', 14),
                   ('F', 14), ('G', 16), ('H', 3)]:
        cw(ws, col, w)

    rh(ws, 1, 8)
    titulo_aba(ws, 2, 2, 7, 'Desempenho por Unidade de Negócio', cliente, sz=14)
    rh(ws, 3, 8)

    # Usa apenas candidatos DigAI para detecção e time_vals
    if '_digai_realizado' in df_d.columns:
        df_c = df_d[df_d['_digai_realizado'] == 'Sim']
    else:
        df_c = df_d

    # Detectar Times — apenas de candidatos DigAI
    if '_time' in df_c.columns:
        time_vals = sorted(set(
            str(t).strip() for t in df_c['_time'].dropna().unique()
            if str(t).strip() not in ('', 'nan')
        ))
    else:
        time_vals = []

    # ── Bloco: segmentação por Time com fórmulas Excel ────────────────────────
    header_bloco(ws, 4, 2, 7, 'Métricas por Time / Unidade de Negócio')
    rh(ws, 5, 22)
    sub_header(ws, 5, [
        (2, 'Time / BU'), (3, 'Total EIs'), (4, 'Aprovados IA'),
        (5, 'Taxa Aprov.'), (6, 'Contratados'), (7, 'Score Médio')
    ])

    has_aprov   = '_aprov'    in df_c.columns and df_c['_aprov'].eq('Sim').any()
    has_score   = '_score_ia' in df_c.columns and df_c['_score_ia'].notna().any()
    has_status  = '_status'   in df_c.columns and df_c['_status'].ne('').any()

    M = "'Base de Dados'!M:M,\"Sim\""

    r = 6
    if time_vals:
        for i, tv in enumerate(time_vals):
            rh(ws, r, 22)
            bg = zebra(i)
            sc(ws, r, 2, tv, bg=bg, bold=True, sz=10)

            # Total EIs DigAI para este time
            cl_tot = sc(ws, r, 3, bg=C['AZUL_CL'], fg=C['AZUL_ESC'], bold=True,
                        ha='center', fmt='#,##0')
            cl_tot.value = f"=COUNTIFS('Base de Dados'!A:A,\"{tv}\",{M})"

            # Aprovados IA
            cl_ap = sc(ws, r, 4, bg=C['AZUL_CL'], fg=C['AZUL_ESC'], ha='center', fmt='#,##0')
            if has_aprov:
                cl_ap.value = (f"=COUNTIFS('Base de Dados'!A:A,\"{tv}\","
                               f"'Base de Dados'!H:H,\"Sim\",{M})")
            else:
                cl_ap.value = '—'
                cl_ap.fill  = PatternFill('solid', start_color=C['CINZA_NA'])

            # Taxa Aprovação
            cl_tx = sc(ws, r, 5, bg=C['VERDE_L'], fg=C['VERDE'], bold=True,
                       ha='center', fmt='0.0%')
            if has_aprov:
                cl_tx.value = (f"=IFERROR(COUNTIFS('Base de Dados'!A:A,\"{tv}\","
                               f"'Base de Dados'!H:H,\"Sim\",{M})"
                               f"/COUNTIFS('Base de Dados'!A:A,\"{tv}\",{M}),0)")
            else:
                cl_tx.value = '—'
                cl_tx.fill  = PatternFill('solid', start_color=C['CINZA_NA'])

            # Contratados DigAI
            cl_ct = sc(ws, r, 6, bg=C['VERDE_L'], fg=C['VERDE'], bold=True,
                       ha='center', fmt='#,##0')
            if has_status:
                cl_ct.value = (f"=COUNTIFS('Base de Dados'!A:A,\"{tv}\","
                               f"'Base de Dados'!F:F,\"Contratado\",{M})")
            else:
                cl_ct.value = '—'
                cl_ct.fill  = PatternFill('solid', start_color=C['CINZA_NA'])

            # Score Médio
            cl_sc = sc(ws, r, 7, bg=C['AZUL_CL'], fg=C['AZUL_ESC'], ha='center', fmt='0.0')
            if has_score:
                cl_sc.value = (f"=IFERROR(AVERAGEIF('Base de Dados'!A:A,\"{tv}\","
                               f"'Base de Dados'!G:G),\"—\")")
            else:
                cl_sc.value = '—'
                cl_sc.fill  = PatternFill('solid', start_color=C['CINZA_NA'])

            r += 1

        # Linha TOTAL
        rh(ws, r, 24)
        sc(ws, r, 2, 'TOTAL GERAL', bg=C['AZUL_ESC'], fg=C['BRANCO'], bold=True)
        cl_tt = sc(ws, r, 3, bg=C['AZUL_ESC'], fg=C['BRANCO'], bold=True, ha='center', fmt='#,##0')
        cl_tt.value = f"=COUNTIF({M})"
        if has_aprov:
            cl_ta = sc(ws, r, 4, bg=C['AZUL_ESC'], fg=C['BRANCO'], bold=True, ha='center', fmt='#,##0')
            cl_ta.value = "=COUNTIF('Base de Dados'!H:H,\"Sim\")"
            cl_tpct = sc(ws, r, 5, bg=C['AZUL_ESC'], fg=C['BRANCO'], bold=True, ha='center', fmt='0.0%')
            cl_tpct.value = f"=IFERROR(COUNTIF('Base de Dados'!H:H,\"Sim\")/COUNTIF({M}),0)"
        else:
            sc(ws, r, 4, '—', bg=C['AZUL_ESC'], fg=C['BRANCO'], ha='center')
            sc(ws, r, 5, '—', bg=C['AZUL_ESC'], fg=C['BRANCO'], ha='center')
        if has_status:
            cl_tct = sc(ws, r, 6, bg=C['AZUL_ESC'], fg=C['BRANCO'], bold=True, ha='center', fmt='#,##0')
            cl_tct.value = f"=COUNTIFS('Base de Dados'!F:F,\"Contratado\",{M})"
        else:
            sc(ws, r, 6, '—', bg=C['AZUL_ESC'], fg=C['BRANCO'], ha='center')
        if has_score:
            cl_tsc = sc(ws, r, 7, bg=C['AZUL_ESC'], fg=C['BRANCO'], bold=True, ha='center', fmt='0.0')
            cl_tsc.value = "=IFERROR(AVERAGE('Base de Dados'!G:G),\"—\")"
        else:
            sc(ws, r, 7, '—', bg=C['AZUL_ESC'], fg=C['BRANCO'], ha='center')
        r += 2

    else:
        # Sem dados de Time — linha única com totais
        rh(ws, r, 22)
        sc(ws, r, 2, 'Todos os Candidatos', bg=C['CINZA'], bold=True)
        cl_t = sc(ws, r, 3, bg=C['AZUL_CL'], fg=C['AZUL_ESC'], bold=True, ha='center', fmt='#,##0')
        cl_t.value = "=COUNTIF('Base de Dados'!M:M,\"Sim\")"
        for c in (4, 5, 6, 7):
            sc(ws, r, c, '— sem campo empresa —', bg=C['CINZA_NA'], fg='595959',
               italic=True, sz=8, ha='center')
        r += 2

    # ── Bloco Observações ─────────────────────────────────────────────────────
    header_bloco(ws, r, 2, 7, 'Como interpretar', bg=C['AZUL_MED'])
    r += 1
    notas = [
        'Time/BU é detectado automaticamente a partir do campo "companyName" da base DigAI.',
        'Taxa de Aprovação = Aprovados IA / Total EIs do Time. Candidatos aprovados avançam para triagem.',
        'Contratados = candidatos com Status Vaga = "Contratado" na base do ATS (Gupy).',
        'Score Médio = média do Score IA dos candidatos daquele Time (coluna G da Base de Dados).',
    ]
    for i, nota in enumerate(notas):
        rh(ws, r, 20)
        mg(ws, r, 2, 7)
        cl = ws.cell(r, 2)
        cl.value     = nota
        cl.font      = Font(name='Calibri', size=9, color='3C3C3C')
        cl.fill      = PatternFill('solid', start_color=zebra(i))
        cl.alignment = Alignment(horizontal='left', vertical='center', indent=2, wrap_text=True)
        brd(cl)
        r += 1


# ── Aba 4 — PIVOT CÁLCULOS ────────────────────────────────────────────────────

def aba_pivot(wb, df_d, cliente):
    ws = wb.create_sheet('Pivot Cálculos')
    ws.sheet_view.showGridLines = False

    for col, w in [('A', 3), ('B', 30), ('C', 16), ('D', 14), ('E', 16), ('F', 14), ('G', 3)]:
        cw(ws, col, w)

    r = 2

    # ── BLOCO A — Score por Faixa ─────────────────────────────────────────────
    header_bloco(ws, r, 2, 6, 'A — Score IA por Faixa')
    r += 1
    sub_header(ws, r, [(2, 'Faixa Score'), (3, 'Qtd Candidatos'), (4, '% do Total'),
                       (5, 'Score Mínimo'), (6, 'Score Máximo')])
    r += 1

    faixas = [
        ('0 – 3  (Baixo)',   '>=0',  '<4',   0,  3),
        ('4 – 6  (Regular)', '>=4',  '<7',   4,  6),
        ('7 – 8  (Bom)',     '>=7',  '<9',   7,  8),
        ('9 – 10 (Excelente)','>=9', '<=10', 9, 10),
    ]
    for i, (label, gte, lt, vmin, vmax) in enumerate(faixas):
        rh(ws, r, 22)
        bg = zebra(i)
        sc(ws, r, 2, label, bg=bg, bold=True)
        cl_q = sc(ws, r, 3, bg=C['AZUL_CL'], fg=C['AZUL_ESC'], bold=True, ha='center', fmt='#,##0')
        cl_q.value = (f"=COUNTIFS('Base de Dados'!G:G,\"{gte}\","
                      f"'Base de Dados'!G:G,\"{lt}\",'Base de Dados'!M:M,\"Sim\")")
        cl_p = sc(ws, r, 4, bg=C['AZUL_CL'], ha='center', fmt='0.0%')
        cl_p.value = (f"=IFERROR(COUNTIFS('Base de Dados'!G:G,\"{gte}\","
                      f"'Base de Dados'!G:G,\"{lt}\",'Base de Dados'!M:M,\"Sim\")"
                      f"/COUNTIF('Base de Dados'!M:M,\"Sim\"),0)")
        sc(ws, r, 5, vmin, bg=bg, ha='center', fmt='0')
        sc(ws, r, 6, vmax, bg=bg, ha='center', fmt='0')
        r += 1

    # Linha Total
    rh(ws, r, 22)
    sc(ws, r, 2, 'TOTAL', bg=C['AZUL_ESC'], fg=C['BRANCO'], bold=True)
    cl_tot = sc(ws, r, 3, bg=C['AZUL_ESC'], fg=C['BRANCO'], bold=True, ha='center', fmt='#,##0')
    cl_tot.value = "=COUNTIF('Base de Dados'!M:M,\"Sim\")"
    sc(ws, r, 4, '100%', bg=C['AZUL_ESC'], fg=C['BRANCO'], bold=True, ha='center')
    sc(ws, r, 5, '—', bg=C['AZUL_ESC'], fg=C['BRANCO'], ha='center')
    sc(ws, r, 6, '—', bg=C['AZUL_ESC'], fg=C['BRANCO'], ha='center')
    r += 2

    # ── BLOCO B — Distribuição por Área ───────────────────────────────────────
    header_bloco(ws, r, 2, 6, 'B — Distribuição por Área')
    r += 1
    sub_header(ws, r, [(2, 'Área'), (3, 'Total EIs'), (4, '% do Total'),
                       (5, 'Aprovados IA'), (6, 'Score Médio')])
    r += 1

    # Usa apenas candidatos DigAI para detecção e Python calculations
    if '_digai_realizado' in df_d.columns:
        df_c = df_d[df_d['_digai_realizado'] == 'Sim']
    else:
        df_c = df_d

    has_aprov = '_aprov'    in df_c.columns and df_c['_aprov'].eq('Sim').any()
    has_score = '_score_ia' in df_c.columns and df_c['_score_ia'].notna().any()

    M = "'Base de Dados'!M:M,\"Sim\""

    if '_area' in df_c.columns:
        areas = sorted(set(
            str(a).strip() for a in df_c['_area'].dropna().unique()
            if str(a).strip() not in ('', 'nan')
        ))[:20]
    else:
        areas = []

    if areas:
        for i, area in enumerate(areas):
            rh(ws, r, 22)
            bg = zebra(i)
            sc(ws, r, 2, area, bg=bg, bold=True)
            cl_q = sc(ws, r, 3, bg=C['AZUL_CL'], fg=C['AZUL_ESC'], ha='center', fmt='#,##0')
            cl_q.value = f"=COUNTIFS('Base de Dados'!B:B,\"{area}\",{M})"
            cl_p = sc(ws, r, 4, bg=C['AZUL_CL'], ha='center', fmt='0.0%')
            cl_p.value = (f"=IFERROR(COUNTIFS('Base de Dados'!B:B,\"{area}\",{M})"
                          f"/COUNTIF({M}),0)")
            cl_ap = sc(ws, r, 5, bg=C['VERDE_L'], fg=C['VERDE'], ha='center', fmt='#,##0')
            if has_aprov:
                cl_ap.value = (f"=COUNTIFS('Base de Dados'!B:B,\"{area}\","
                               f"'Base de Dados'!H:H,\"Sim\")")
            else:
                cl_ap.value = '—'
                cl_ap.fill  = PatternFill('solid', start_color=C['CINZA_NA'])
            cl_sc = sc(ws, r, 6, bg=C['AZUL_CL'], fg=C['AZUL_ESC'], ha='center', fmt='0.0')
            if has_score:
                cl_sc.value = (f"=IFERROR(AVERAGEIF('Base de Dados'!B:B,\"{area}\","
                               f"'Base de Dados'!G:G),\"—\")")
            else:
                cl_sc.value = '—'
                cl_sc.fill  = PatternFill('solid', start_color=C['CINZA_NA'])
            r += 1
    else:
        rh(ws, r, 22)
        mg(ws, r, 2, 6)
        cl = ws.cell(r, 2)
        cl.value     = 'Campo Área não detectado na base enviada (coluna workspace/area_cand)'
        cl.font      = Font(name='Calibri', size=9, italic=True, color='595959')
        cl.fill      = PatternFill('solid', start_color=C['CINZA_NA'])
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)
        r += 1

    r += 1

    # ── BLOCO C — Evolução Mensal ──────────────────────────────────────────────
    header_bloco(ws, r, 2, 6, 'C — Evolução Mensal de Entrevistas')
    r += 1
    sub_header(ws, r, [(2, 'Mês / Período'), (3, 'Entrevistas'), (4, 'Aprovados IA'),
                       (5, 'Taxa Aprovação'), (6, 'Score Médio')])
    r += 1

    if '_data_aplic' in df_c.columns and df_c['_data_aplic'].notna().any():
        tmp = df_c[df_c['_data_aplic'].notna()].copy()
        tmp['_mes'] = pd.to_datetime(tmp['_data_aplic'], errors='coerce').dt.to_period('M')
        mensal = (tmp.groupby('_mes')
                  .agg(
                      total=('_email', 'count'),
                      aprov=('_aprov', lambda x: (x == 'Sim').sum()) if '_aprov' in tmp.columns else ('_email', 'count'),
                      score=('_score_ia', 'mean') if '_score_ia' in tmp.columns else ('_email', 'count'),
                  )
                  .reset_index()
                  .sort_values('_mes'))

        for i, row_m in mensal.iterrows():
            rh(ws, r, 22)
            bg = zebra(int(i) % 2)
            mes_label = str(row_m['_mes'])
            total     = int(row_m['total'])
            aprov     = int(row_m.get('aprov', 0)) if has_aprov else None
            score_m   = float(row_m.get('score', np.nan)) if has_score else None

            sc(ws, r, 2, mes_label, bg=bg, bold=True)
            sc(ws, r, 3, total, bg=C['AZUL_CL'], fg=C['AZUL_ESC'], ha='center', fmt='#,##0')
            if aprov is not None:
                sc(ws, r, 4, aprov, bg=C['VERDE_L'], fg=C['VERDE'], ha='center', fmt='#,##0')
                taxa = aprov / total if total else 0
                sc(ws, r, 5, taxa, bg=C['VERDE_L'], fg=C['VERDE'], ha='center', fmt='0.0%')
            else:
                sc(ws, r, 4, '—', bg=C['CINZA_NA'], ha='center')
                sc(ws, r, 5, '—', bg=C['CINZA_NA'], ha='center')
            if score_m is not None and not np.isnan(score_m):
                sc(ws, r, 6, round(score_m, 1), bg=C['AZUL_CL'], fg=C['AZUL_ESC'],
                   ha='center', fmt='0.0')
            else:
                sc(ws, r, 6, '—', bg=C['CINZA_NA'], ha='center')
            r += 1
    else:
        rh(ws, r, 22)
        mg(ws, r, 2, 6)
        cl = ws.cell(r, 2)
        cl.value     = 'Campo Data de Aplicação não encontrado na base enviada'
        cl.font      = Font(name='Calibri', size=9, italic=True, color='595959')
        cl.fill      = PatternFill('solid', start_color=C['CINZA_NA'])
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)
        r += 1


# ── Aba 5 — BASE DE DADOS ─────────────────────────────────────────────────────

def aba_base_dados(wb, df_d):
    """
    13 colunas fixas — as fórmulas das outras abas dependem desta ordem exata:
      A: Time          B: Área          C: Nome          D: Email
      E: Vaga          F: Status Vaga   G: Score IA      H: Aprovado IA
      I: Score Editado J: Req. Atendido K: Dias Vaga Aberta L: Data Aplicação
      M: DigAI Realizado  ← filtro chave para fórmulas COUNTIF(M:M,"Sim")
    """
    ws = wb.create_sheet('Base de Dados')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    COLUNAS = [
        ('Time',              16),
        ('Área',              22),
        ('Nome',              28),
        ('Email',             32),
        ('Vaga',              36),
        ('Status Vaga',       16),
        ('Score IA',          12),
        ('Aprovado IA',       14),
        ('Score Editado',     14),
        ('Req. Atendido',     14),
        ('Dias Vaga Aberta',  16),
        ('Data Aplicação',    16),
        ('DigAI Realizado',   16),
    ]

    for i, (_, w) in enumerate(COLUNAS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    rh(ws, 1, 28)
    for ci, (nome, _) in enumerate(COLUNAS, 1):
        cl = ws.cell(1, ci, nome)
        cl.font      = Font(name='Calibri', size=9, bold=True, color=C['BRANCO'])
        cl.fill      = PatternFill('solid', start_color=C['AZUL_ESC'])
        cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        brd(cl)

    if df_d is None or len(df_d) == 0:
        return

    # Mapeamento interno → coluna Base de Dados (ordem = A..M)
    INT_COLS = [
        '_time', '_area', '_nome', '_email', '_vaga', '_status',
        '_score_ia', '_aprov', '_score_ed', '_req', '_dias_vaga', '_data_aplic',
        '_digai_realizado',
    ]

    ROW_CAP  = 10_000
    df_slice = df_d.head(ROW_CAP)

    for row_data in df_slice.to_dict('records'):
        row_vals = []
        for ic in INT_COLS:
            val = row_data.get(ic, '')
            if val is None or (isinstance(val, float) and str(val) == 'nan'):
                row_vals.append('')
            elif hasattr(val, 'strftime'):
                try:
                    row_vals.append(val.strftime('%d/%m/%Y'))
                except (ValueError, TypeError):
                    row_vals.append('')
            elif isinstance(val, (np.integer,)):
                row_vals.append(int(val))
            elif isinstance(val, (np.floating,)):
                row_vals.append(float(val) if not np.isnan(val) else '')
            else:
                row_vals.append(val)
        ws.append(row_vals)

    # Formatação Score IA e Score Editado (colunas G e I = col 7 e 9)
    # Coluna M (13) DigAI Realizado: Sim=verde, Não=cinza
    for row_idx in range(2, min(len(df_slice) + 2, 10002)):
        for col_idx in (7, 9):
            ws.cell(row_idx, col_idx).number_format = '0.0'
        ws.cell(row_idx, 12).number_format = 'DD/MM/YYYY'
        cl_m = ws.cell(row_idx, 13)
        cl_m.alignment = Alignment(horizontal='center', vertical='center')
        if cl_m.value == 'Sim':
            cl_m.font = Font(name='Calibri', size=9, bold=True, color=C['VERDE'])
            cl_m.fill = PatternFill('solid', start_color=C['VERDE_L'])
        else:
            cl_m.font = Font(name='Calibri', size=9, color='595959')
            cl_m.fill = PatternFill('solid', start_color=C['CINZA_NA'])

    if len(df_d) > ROW_CAP:
        note_row = ROW_CAP + 3
        cl = ws.cell(note_row, 1,
                     f'Mostrando {ROW_CAP:,} de {len(df_d):,} registros. '
                     'Exporte o CSV completo para análise integral.')
        cl.font = Font(name='Calibri', size=9, italic=True, color='595959')


# ── Aba 6 (opt) — RANKING POR EMPRESA ────────────────────────────────────────

def aba_ranking_empresa(wb, df_d, cliente):
    ws = wb.create_sheet('Ranking por Empresa')
    ws.sheet_view.showGridLines = False

    for col, w in [('A', 3), ('B', 30), ('C', 14), ('D', 14), ('E', 14),
                   ('F', 14), ('G', 14), ('H', 3)]:
        cw(ws, col, w)

    rh(ws, 1, 8)
    titulo_aba(ws, 2, 2, 7, 'Ranking por Empresa / companyName', cliente, sz=14)
    rh(ws, 3, 8)

    has_aprov  = '_aprov'    in df_d.columns and df_d['_aprov'].ne('').any()
    has_score  = '_score_ia' in df_d.columns and df_d['_score_ia'].notna().any()
    has_status = '_status'   in df_d.columns and df_d['_status'].ne('').any()

    # Calcular ranking em Python para dados dinâmicos
    if '_time' in df_d.columns:
        grp = df_d.groupby('_time')
        empresas_data = []
        for empresa, g in grp:
            if not str(empresa).strip() or str(empresa).strip() == 'nan':
                continue
            total  = len(g)
            aprov  = int((g['_aprov'] == 'Sim').sum()) if has_aprov else None
            cont   = int((g['_status'].str.lower() == 'contratado').sum()) if has_status else None
            sc_med = float(g['_score_ia'].mean()) if has_score and g['_score_ia'].notna().any() else None
            empresas_data.append((empresa, total, aprov, cont, sc_med))

        empresas_data.sort(key=lambda x: x[1], reverse=True)
    else:
        empresas_data = []

    header_bloco(ws, 4, 2, 7, 'Empresas rankeadas por volume de Entrevistas DigAI')
    sub_header(ws, 5, [
        (2, '#'), (3, 'Empresa / BU'), (4, 'Total EIs'), (5, 'Aprovados'),
        (6, 'Contratados'), (7, 'Score Médio')
    ])

    r = 6
    for rank, (emp, total, aprov, cont, sc_med) in enumerate(empresas_data, 1):
        rh(ws, r, 22)
        bg = zebra(rank)
        sc(ws, r, 2, rank, bg=bg, ha='center', bold=True)
        sc(ws, r, 3, str(emp), bg=bg, bold=(rank == 1))
        sc(ws, r, 4, total, bg=C['AZUL_CL'], fg=C['AZUL_ESC'], bold=True, ha='center', fmt='#,##0')
        if aprov is not None:
            sc(ws, r, 5, aprov, bg=C['VERDE_L'], fg=C['VERDE'], ha='center', fmt='#,##0')
        else:
            sc(ws, r, 5, '—', bg=C['CINZA_NA'], ha='center')
        if cont is not None:
            sc(ws, r, 6, cont, bg=C['VERDE_L'], fg=C['VERDE'], ha='center', fmt='#,##0')
        else:
            sc(ws, r, 6, '—', bg=C['CINZA_NA'], ha='center')
        if sc_med is not None and not np.isnan(sc_med):
            sc(ws, r, 7, round(sc_med, 1), bg=C['AZUL_CL'], fg=C['AZUL_ESC'],
               ha='center', fmt='0.0')
        else:
            sc(ws, r, 7, '—', bg=C['CINZA_NA'], ha='center')
        r += 1

    if not empresas_data:
        rh(ws, r, 22)
        mg(ws, r, 2, 7)
        cl = ws.cell(r, 2)
        cl.value     = 'Campo companyName não detectado na base DigAI enviada'
        cl.font      = Font(name='Calibri', size=9, italic=True, color='595959')
        cl.fill      = PatternFill('solid', start_color=C['CINZA_NA'])
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)


# ── Aba 7 (opt) — ORIGEM DOS CANDIDATOS ──────────────────────────────────────

def aba_origem_candidatos(wb, df_d, cliente):
    ws = wb.create_sheet('Origem dos Candidatos')
    ws.sheet_view.showGridLines = False

    for col, w in [('A', 3), ('B', 22), ('C', 24), ('D', 14), ('E', 14), ('F', 14), ('G', 3)]:
        cw(ws, col, w)

    rh(ws, 1, 8)
    titulo_aba(ws, 2, 2, 6, 'Origem Geográfica dos Candidatos (por DDD)', cliente, sz=14)
    rh(ws, 3, 8)

    def _ddd_to_estado(phone_str):
        digits = re.sub(r'\D', '', str(phone_str))
        if len(digits) >= 10:
            ddd = digits[:2] if digits[0] != '0' else digits[1:3]
            return DDD_ESTADO.get(ddd[:2], 'Outro')
        return 'Sem DDD'

    if 'phone' in df_d.columns or 'phone_raw' in df_d.columns:
        phone_col = 'phone' if 'phone' in df_d.columns else 'phone_raw'
        tmp = df_d[df_d[phone_col].notna() & df_d[phone_col].ne('')].copy()
        tmp['_estado'] = tmp[phone_col].astype(str).apply(_ddd_to_estado)
    else:
        tmp = pd.DataFrame()

    header_bloco(ws, 4, 2, 6, 'Candidatos por Estado (extraído do DDD do celular)')
    sub_header(ws, 5, [
        (2, 'UF'), (3, 'Estado'), (4, 'Total'), (5, '% do Total'), (6, 'Aprovados IA')
    ])

    r = 6
    if len(tmp) > 0:
        grp = (tmp.groupby('_estado')
               .agg(total=('_email', 'count'),
                    aprov=('_aprov', lambda x: (x == 'Sim').sum())
                    if '_aprov' in tmp.columns else ('_email', 'count'))
               .reset_index()
               .sort_values('total', ascending=False))

        total_geral = int(grp['total'].sum())
        has_aprov_o = '_aprov' in tmp.columns and tmp['_aprov'].ne('').any()

        for i, row_g in grp.iterrows():
            uf    = str(row_g['_estado'])
            nome  = ESTADO_NOME.get(uf, uf)
            total = int(row_g['total'])
            aprov = int(row_g.get('aprov', 0)) if has_aprov_o else None
            pct   = total / total_geral if total_geral else 0

            rh(ws, r, 22)
            bg = zebra(int(i) % 2)
            sc(ws, r, 2, uf, bg=bg, bold=True, ha='center')
            sc(ws, r, 3, nome, bg=bg)
            sc(ws, r, 4, total, bg=C['AZUL_CL'], fg=C['AZUL_ESC'], ha='center', fmt='#,##0')
            sc(ws, r, 5, pct, bg=C['AZUL_CL'], ha='center', fmt='0.0%')
            if aprov is not None:
                sc(ws, r, 6, aprov, bg=C['VERDE_L'], fg=C['VERDE'], ha='center', fmt='#,##0')
            else:
                sc(ws, r, 6, '—', bg=C['CINZA_NA'], ha='center')
            r += 1
    else:
        rh(ws, r, 22)
        mg(ws, r, 2, 6)
        cl = ws.cell(r, 2)
        cl.value     = 'Campo phoneNumber não disponível (cobertura < 50%)'
        cl.font      = Font(name='Calibri', size=9, italic=True, color='595959')
        cl.fill      = PatternFill('solid', start_color=C['CINZA_NA'])
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)


# ── FUNÇÃO PRINCIPAL ──────────────────────────────────────────────────────────

def gerar_relatorio(df_base=None, params=None, output_path='relatorio.xlsx',
                    cliente='Cliente', periodo='',
                    # Parâmetros legados (ignorados — mantidos para compatibilidade)
                    kpis=None, funil=None, tempo_etapas=None) -> str:
    """
    Gera o relatório DigAI com 5 abas obrigatórias + 2 opcionais.

    Args:
        df_base:      DataFrame unificado (filtragem para DigAI é feita internamente)
        params:       Dict com mensalidade, salario_clt, prod_max, etc.
        output_path:  Caminho do .xlsx de saída
        cliente:      Nome do cliente
        periodo:      Período do relatório (ex: "Jan/2025 a Mar/2025")

    Returns:
        output_path
    """
    params = params or {}

    # Construir df normalizado com 13 colunas internas (inclui todos ATS)
    print('[XLS] Iniciando gerar_relatorio...', flush=True)
    df_d = _build_base_df(df_base)
    print(f'[XLS] _build_base_df: {len(df_d)} linhas, colunas={list(df_d.columns)}', flush=True)

    # df_c = apenas candidatos DigAI (para abas opcionais e cálculos Python)
    if '_digai_realizado' in df_d.columns:
        df_c = df_d[df_d['_digai_realizado'] == 'Sim']
    else:
        df_c = df_d
    print(f'[XLS] df_c (DigAI): {len(df_c)} linhas', flush=True)

    wb = Workbook()

    # Tab 1: Indicadores DigAI
    print('[XLS] aba_indicadores...', flush=True)
    aba_indicadores(wb, df_d, params, cliente, periodo)

    # Tab 2: Calculadora ROI
    print('[XLS] aba_calculadora...', flush=True)
    aba_calculadora(wb, params, cliente)

    # Tab 3: Por Time
    print('[XLS] aba_por_time...', flush=True)
    aba_por_time(wb, df_d, cliente)

    # Tab 4: Pivot Cálculos
    print('[XLS] aba_pivot...', flush=True)
    aba_pivot(wb, df_d, cliente)

    # Tab 5: Base de Dados (criada por último na ordem visual)
    print('[XLS] aba_base_dados...', flush=True)
    aba_base_dados(wb, df_d)

    # Abas opcionais — baseadas em df_c (apenas DigAI)
    if '_time' in df_c.columns:
        empresas_validas = [
            t for t in df_c['_time'].dropna().unique()
            if str(t).strip() not in ('', 'nan')
        ]
        if len(set(str(e).strip() for e in empresas_validas)) >= 3:
            print('[XLS] aba_ranking_empresa...', flush=True)
            aba_ranking_empresa(wb, df_c, cliente)

    phone_col = next((c for c in ('phone', 'phone_raw') if c in df_c.columns), None)
    if phone_col:
        total_rows = len(df_c)
        phone_cov  = df_c[phone_col].notna().sum() / total_rows if total_rows else 0
        non_empty  = df_c[phone_col].ne('').sum() / total_rows if total_rows else 0
        if max(phone_cov, non_empty) >= 0.50:
            print('[XLS] aba_origem_candidatos...', flush=True)
            aba_origem_candidatos(wb, df_c, cliente)

    print(f'[XLS] wb.save({output_path})...', flush=True)
    wb.save(output_path)
    print(f'✅ Relatório gerado: {output_path}')
    print(f'   Abas: {wb.sheetnames}')
    print(f'   Total ATS na Base: {len(df_d):,} | DigAI Realizado: {len(df_c):,}')
    return output_path
